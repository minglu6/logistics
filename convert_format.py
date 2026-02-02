#!/usr/bin/env python3
"""
将临努对账单转换为模板格式
"""
import openpyxl
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from copy import copy
import re

def parse_store_info(store_text):
    """
    解析店名信息，从"店名：公里数，店名：公里数"格式中提取店名
    返回店名列表（用换行符连接）和总公里数
    """
    if not store_text or store_text == '':
        return '', 0

    # 分割每个店铺信息
    stores = store_text.split('，')
    store_names = []
    total_km = 0

    for store in stores:
        # 使用中文冒号分割店名和公里数
        parts = store.split('：')
        if len(parts) >= 2:
            store_name = parts[0].strip()
            km_str = parts[1].strip()
            store_names.append(store_name)

            # 提取公里数
            try:
                km = float(km_str)
                total_km += km
            except:
                pass

    return '\n'.join(store_names), total_km

def copy_cell_style(source_cell, target_cell):
    """复制单元格样式"""
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def convert_to_template_format(source_file, output_file, template_file):
    """转换文件格式"""
    # 读取源文件
    wb_source = openpyxl.load_workbook(source_file)
    ws_source = wb_source.active

    # 读取模板文件以复制样式
    wb_template = openpyxl.load_workbook(template_file)
    ws_template = wb_template.active

    # 创建新工作簿
    wb_output = openpyxl.Workbook()
    ws_output = wb_output.active
    ws_output.title = 'Sheet1 (3)'

    # 写入表头
    headers = ['序号', '日期', '店名', '公里数', '公里数',
               '不含税单价', '含税单价', '不含税合价（运费）', '含税合价（运费）',
               '司机价格', '司机价格', '司机姓名', '照片', '备注']
    ws_output.append(headers)

    # 复制表头样式
    for col_idx in range(1, 15):
        template_cell = ws_template.cell(1, col_idx)
        output_cell = ws_output.cell(1, col_idx)
        copy_cell_style(template_cell, output_cell)

    # 遍历源文件数据（从第2行开始，跳过表头）
    seq_num = 1
    for row_idx in range(2, ws_source.max_row + 1):
        # 读取源数据
        date_val = ws_source.cell(row_idx, 1).value  # 日期
        region = ws_source.cell(row_idx, 2).value  # 区域
        store_text = ws_source.cell(row_idx, 3).value  # 门店
        km_formula = ws_source.cell(row_idx, 4).value  # 公里数（公式）
        price = ws_source.cell(row_idx, 5).value  # 单价（含税）
        freight = ws_source.cell(row_idx, 6).value  # 运费
        remark = ws_source.cell(row_idx, 7).value  # 备注

        # 如果没有店铺信息，跳过
        if not store_text:
            continue

        # 解析店名信息
        store_names, total_km = parse_store_info(store_text)

        # 创建新行数据
        new_row = [
            seq_num,  # 序号
            date_val,  # 日期
            store_names,  # 店名（换行符分隔）
            km_formula,  # 公里数（保持原公式）
            total_km if total_km > 0 else None,  # 公里数（计算值）
            None,  # 不含税单价（使用公式）
            price,  # 含税单价
            None,  # 不含税合价（使用公式）
            None,  # 含税合价（使用公式）
            None,  # 司机价格单价（使用公式）
            None,  # 司机价格合价（使用公式）
            None,  # 司机姓名
            None,  # 照片
            remark  # 备注
        ]

        ws_output.append(new_row)

        # 当前行号（从2开始，因为第1行是表头）
        current_row = seq_num + 1

        # 设置公式
        # 不含税单价 = 含税单价 / 1.09
        ws_output.cell(current_row, 6).value = f'=G{current_row}/1.09'

        # 含税单价（如果来源文件是公式，使用阶梯价格公式）
        if isinstance(price, (int, float)):
            ws_output.cell(current_row, 7).value = price
        else:
            ws_output.cell(current_row, 7).value = f'=IF(D{current_row}<=100,440,IF(D{current_row}<=200,4.2,IF(D{current_row}<=300,4,3.9)))'

        # 不含税合价 = 公里数 * 不含税单价
        ws_output.cell(current_row, 8).value = f'=D{current_row}*F{current_row}'

        # 含税合价 = 公里数 * 含税单价
        ws_output.cell(current_row, 9).value = f'=D{current_row}*G{current_row}'

        # 司机价格单价
        ws_output.cell(current_row, 10).value = f'=IF(D{current_row}<=100,400,IF(D{current_row}<=300,3.2,3))'

        # 司机价格合价 = 公里数 * 司机价格单价
        ws_output.cell(current_row, 11).value = f'=D{current_row}*J{current_row}'

        # 复制模板行的样式到当前行
        template_row = 2  # 使用模板的第2行作为样式参考
        for col_idx in range(1, 15):
            template_cell = ws_template.cell(template_row, col_idx)
            output_cell = ws_output.cell(current_row, col_idx)
            copy_cell_style(template_cell, output_cell)

        # 确保店名单元格自动换行
        cell = ws_output.cell(current_row, 3)
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='general')

        seq_num += 1

    # 复制列宽
    for col_idx in range(1, 15):
        col_letter_template = openpyxl.utils.get_column_letter(col_idx)
        col_letter_output = openpyxl.utils.get_column_letter(col_idx)
        if ws_template.column_dimensions[col_letter_template].width:
            ws_output.column_dimensions[col_letter_output].width = ws_template.column_dimensions[col_letter_template].width

    # 保存输出文件
    wb_output.save(output_file)
    wb_template.close()
    wb_source.close()
    wb_output.close()

    print(f"转换完成！")
    print(f"源文件: {source_file}")
    print(f"输出文件: {output_file}")
    print(f"共转换 {seq_num - 1} 行数据")

if __name__ == '__main__':
    template_file = 'data/jiangxi/summary/惠宜选江西仓对账单模板.xlsx'
    source_file = 'data/jiangxi/summary/2026/惠宜选物流对账单--临努--1月(2).xlsx'
    output_file = 'data/jiangxi/summary/2026/惠宜选物流对账单--临努--1月(2)_转换后.xlsx'

    convert_to_template_format(source_file, output_file, template_file)
