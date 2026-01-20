#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据分析工具
支持多种分析模式：结构分析、距离分析、数据预览等

用法:
    python -m scripts.analysis.analyze_excel --mode structure --input 对账单.xlsx
    python -m scripts.analysis.analyze_excel --mode distances --input 对账单.xlsx
    python -m scripts.analysis.analyze_excel --mode preview --input 对账单.xlsx --rows 10
"""

import argparse
import os
import re
import sys

import openpyxl
import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.utils.common import parse_shop_and_distance


def analyze_structure(excel_file, rows=10):
    """
    分析Excel结构
    """
    print("=" * 80)
    print("Excel结构分析")
    print("=" * 80)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    print(f"\n工作表: {ws.title}")
    print(f"最大行数: {ws.max_row}")
    print(f"最大列数: {ws.max_column}")

    # 显示列标题
    print("\n列标题（第1行）:")
    for col in range(1, min(ws.max_column + 1, 10)):
        cell_value = ws.cell(row=1, column=col).value
        print(f"  列{col}: {cell_value}")

    # 显示前几行数据
    print(f"\n前{rows}行数据预览:")
    for row_idx in range(2, min(rows + 2, ws.max_row + 1)):
        print(f"\n--- 第{row_idx}行 ---")
        for col in range(1, min(ws.max_column + 1, 6)):
            cell_value = ws.cell(row=row_idx, column=col).value
            if cell_value:
                value_str = str(cell_value)[:100]
                if len(str(cell_value)) > 100:
                    value_str += "..."
                print(f"  列{col}: {value_str}")

    wb.close()


def analyze_distances(excel_file):
    """
    分析对账单中的距离数据
    """
    print("=" * 80)
    print("距离数据分析")
    print("=" * 80)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    total_shops = 0
    shops_with_distance = 0
    shops_without_distance = 0
    distance_values = []

    for row_idx in range(2, ws.max_row + 1):
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if not shop_names_cell:
            continue

        shop_lines = str(shop_names_cell).split('\n')
        for line in shop_lines:
            line = line.strip()
            if not line:
                continue

            total_shops += 1
            shop_name, distance = parse_shop_and_distance(line)

            if distance is not None:
                shops_with_distance += 1
                distance_values.append(distance)
            elif '-?km' in line:
                shops_without_distance += 1

    print(f"\n统计结果:")
    print(f"  总店铺数: {total_shops}")
    print(f"  有距离: {shops_with_distance} ({shops_with_distance/total_shops*100:.1f}%)")
    print(f"  无距离: {shops_without_distance} ({shops_without_distance/total_shops*100:.1f}%)")

    if distance_values:
        print(f"\n距离分布:")
        print(f"  最小: {min(distance_values)} km")
        print(f"  最大: {max(distance_values)} km")
        print(f"  平均: {sum(distance_values)/len(distance_values):.1f} km")

    wb.close()


def preview_data(excel_file, rows=10):
    """
    预览对账单数据
    """
    print("=" * 80)
    print("数据预览")
    print("=" * 80)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    for row_idx in range(2, min(rows + 2, ws.max_row + 1)):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        print(f"\n车次 {vehicle_no} (日期: {date_value}):")
        print("-" * 60)

        if shop_names_cell:
            shop_lines = str(shop_names_cell).split('\n')
            for i, line in enumerate(shop_lines, 1):
                line = line.strip()
                if not line:
                    continue

                # 检查距离状态
                if '-?km' in line:
                    status = "待填充"
                elif re.search(r'-\d+(\.\d+)?km', line):
                    status = "已填充"
                else:
                    status = "无距离"

                print(f"  {i}. {line} [{status}]")

    wb.close()


def check_merged_cells(excel_file):
    """
    检查合并单元格
    """
    print("=" * 80)
    print("合并单元格分析")
    print("=" * 80)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    merged_ranges = list(ws.merged_cells.ranges)
    print(f"\n合并单元格数量: {len(merged_ranges)}")

    if merged_ranges:
        print("\n前20个合并区域:")
        for i, merged_range in enumerate(merged_ranges[:20], 1):
            print(f"  {i}. {merged_range}")

    wb.close()


def main():
    parser = argparse.ArgumentParser(description='Excel数据分析工具')
    parser.add_argument('--mode', '-m', required=True,
                        choices=['structure', 'distances', 'preview', 'merged'],
                        help='分析模式: structure/distances/preview/merged')
    parser.add_argument('--input', '-i', required=True,
                        help='输入Excel文件路径')
    parser.add_argument('--rows', '-r', type=int, default=10,
                        help='预览行数（默认10）')

    args = parser.parse_args()

    if args.mode == 'structure':
        analyze_structure(args.input, args.rows)
    elif args.mode == 'distances':
        analyze_distances(args.input)
    elif args.mode == 'preview':
        preview_data(args.input, args.rows)
    elif args.mode == 'merged':
        check_merged_cells(args.input)


if __name__ == '__main__':
    main()
