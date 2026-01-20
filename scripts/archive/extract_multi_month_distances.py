#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从9月、10月、11月、12月对账单中提取距离数据
按照时间顺序处理，如果有冲突且距离相差5km内，以日期新的为准
"""

import pandas as pd
from openpyxl import load_workbook
import json
import re
from collections import defaultdict

STARTING_POINT = "丰树合肥现代综合产业园"
CONFLICT_THRESHOLD = 5.0  # 5km内视为相同路段

def parse_shop_and_distance(shop_line):
    """解析店名行，提取店名和距离"""
    shop_line = shop_line.strip()
    pattern = r'^(.+?)-(\d+(?:\.\d+)?)km$'
    match = re.match(pattern, shop_line)

    if match:
        shop_name = match.group(1)
        distance = float(match.group(2))
        return shop_name, distance
    else:
        return shop_line, None


def extract_routes_from_excel(file_path, has_distance_in_name=False):
    """从对账单Excel中提取路线数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    routes = []

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        shops = []
        distances = []

        if shop_names_cell:
            shop_lines = [s.strip() for s in str(shop_names_cell).split('\n') if s.strip()]

            for shop_line in shop_lines:
                if has_distance_in_name:
                    shop_name, distance = parse_shop_and_distance(shop_line)
                    shops.append(shop_name)
                    distances.append(distance)
                else:
                    shops.append(shop_line)
                    distances.append(None)

        route = {
            'vehicle_no': vehicle_no,
            'date': date_value,
            'shops': shops,
            'distances': distances
        }

        routes.append(route)

    return routes


def build_segments(route):
    """根据README规则构建路段"""
    segments = []
    shops = route['shops']
    distances = route['distances']

    if not shops:
        return segments

    first_distance = distances[0] if len(distances) > 0 else None
    segments.append((STARTING_POINT, shops[0], first_distance))

    for i in range(1, len(shops)):
        distance = distances[i] if i < len(distances) else None
        segments.append((shops[i-1], shops[i], distance))

    return segments


print("从多个月份提取距离数据...")
print("="*100)

# 定义月份文件（按时间顺序）
month_files = [
    ('9月', r'D:\Work\logistics\data\summary\惠宜选合肥仓9月份对账单9.22.xlsx'),
    ('10月', r'D:\Work\logistics\data\summary\惠宜选合肥仓10月份对账单(2).xlsx'),
    ('11月', r'D:\Work\logistics\data\summary\惠宜选合肥仓11月份对账单-完整版.xlsx'),
    ('12月', r'D:\Work\logistics\data\summary\惠宜选合肥仓12月份对账单.xlsx'),
]

# 存储每个月份的距离数据
all_month_data = {}
conflicts = []
large_differences = []

for month_name, file_path in month_files:
    print(f"\n处理 {month_name} 数据...")
    routes = extract_routes_from_excel(file_path, has_distance_in_name=True)

    # 构建本月份的距离字典
    month_distance_map = {}

    for route in routes:
        segments = build_segments(route)
        for seg in segments:
            key = (seg[0], seg[1])
            distance = seg[2]

            if distance is not None:
                if key not in month_distance_map:
                    month_distance_map[key] = []
                month_distance_map[key].append(distance)

    # 计算平均距离
    month_avg_distances = {}
    for key, distances in month_distance_map.items():
        avg_distance = sum(distances) / len(distances)
        month_avg_distances[key] = round(avg_distance, 2)

    all_month_data[month_name] = month_avg_distances
    print(f"  提取到 {len(month_avg_distances)} 个路段")

# 合并数据（按时间顺序，新数据覆盖旧数据，但要检查冲突）
print("\n" + "="*100)
print("合并距离数据...")
final_distances = {}
source_month = {}  # 记录每个路段来自哪个月份

for month_name, month_distances in all_month_data.items():
    for key, new_distance in month_distances.items():
        segment_key = f"{key[0]} -> {key[1]}"

        if segment_key in final_distances:
            old_distance = final_distances[segment_key]
            diff = abs(new_distance - old_distance)

            if diff <= CONFLICT_THRESHOLD:
                # 距离相差在5km内，使用新数据
                conflicts.append({
                    'segment': segment_key,
                    'old_distance': old_distance,
                    'new_distance': new_distance,
                    'difference': diff,
                    'old_month': source_month[segment_key],
                    'new_month': month_name
                })
                final_distances[segment_key] = new_distance
                source_month[segment_key] = month_name
            else:
                # 距离相差超过5km，记录但保留旧数据（不更新）
                large_differences.append({
                    'segment': segment_key,
                    'old_distance': old_distance,
                    'new_distance': new_distance,
                    'difference': diff,
                    'old_month': source_month[segment_key],
                    'new_month': month_name,
                    'used_distance': old_distance  # 使用旧数据
                })
                # 不更新 final_distances 和 source_month，保留旧数据
        else:
            final_distances[segment_key] = new_distance
            source_month[segment_key] = month_name

# 保存主JSON文件
output_file = r'D:\Work\logistics\data\cache\reusable_distances.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(final_distances, f, ensure_ascii=False, indent=2)

print(f"\n✓ JSON文件已保存: {output_file}")
print(f"✓ 总计 {len(final_distances)} 个路段的距离数据")

# 保存大差异数据到单独的JSON文件
if large_differences:
    large_diff_output = r'D:\Work\logistics\data\cache\large_distance_differences.json'
    large_diff_data = []

    for d in large_differences:
        large_diff_data.append({
            'segment': d['segment'],
            'old_month': d['old_month'],
            'old_distance_km': d['old_distance'],
            'new_month': d['new_month'],
            'new_distance_km': d['new_distance'],
            'difference_km': round(d['difference'], 2),
            'used_distance_km': d['used_distance'],
            'warning': '距离差异超过5km，已保留旧数据，请人工检查后决定使用哪个'
        })

    with open(large_diff_output, 'w', encoding='utf-8') as f:
        json.dump(large_diff_data, f, ensure_ascii=False, indent=2)

    print(f"\n✓ 大差异数据已保存: {large_diff_output}")
    print(f"✓ 包含 {len(large_diff_data)} 个需要检查的路段")

# 统计报告
print("\n" + "="*100)
print("数据来源统计")
print("="*100)
for month_name in ['9月', '10月', '11月', '12月']:
    count = sum(1 for m in source_month.values() if m == month_name)
    print(f"{month_name}: {count} 个路段")

# 冲突报告
if conflicts:
    print("\n" + "="*100)
    print(f"发现 {len(conflicts)} 个冲突路段（距离相差5km内，已使用新数据）")
    print("="*100)
    for c in conflicts:
        print(f"\n路段: {c['segment']}")
        print(f"  {c['old_month']}: {c['old_distance']} km")
        print(f"  {c['new_month']}: {c['new_distance']} km (使用)")
        print(f"  差值: {c['difference']:.2f} km")

# 大差异报告
if large_differences:
    print("\n" + "="*100)
    print(f"警告：发现 {len(large_differences)} 个路段距离相差超过5km，已保留旧数据")
    print("="*100)
    for d in large_differences:
        print(f"\n路段: {d['segment']}")
        print(f"  {d['old_month']}: {d['old_distance']} km (保留)")
        print(f"  {d['new_month']}: {d['new_distance']} km")
        print(f"  差值: {d['difference']:.2f} km ⚠️")
    print("\n这些路段已保留旧数据，请人工检查后决定使用哪个数据！")

print("\n" + "="*100)
print("提取完成！")
print("="*100)
