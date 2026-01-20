#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Fill shop names into Jiangxi billing Excel file for Jan 1st.
"""

import openpyxl
import os

def main():
    file_path = r'D:\Work\logistics\data\jiangxi\summary\惠宜选江西仓1月份对账单.xlsx'
    
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        print(f"Loaded workbook: {file_path}")
    except Exception as e:
        print(f"Failed to load workbook: {e}")
        return

    # Data to fill
    # Format: {excel_row_number: shop_names_string}
    # Note: Vehicle 1 is already in Row 2 (as seen in analysis).
    # We need to fill Vehicle 2 (Row 3), Vehicle 3 (Row 4), Vehicle 4 (Row 5).
    
    updates = {
        3: [
            "共橙一站式超市（吉水万里大道店）",
            "共橙一站式超市（吉安县赣江大道店）",
            "厉臣省钱超市（泰和钦顺路店）",
            "共橙一站式超市（泰和澄江大道店）",
            "共橙一站式超市（永新湘赣大道店）",
            "共橙一站式超市（井冈山步云山路店）",
            "共橙一站式超市（万安建设路店）",
            "共橙一站式超市（遂川川江南路店）"
        ],
        4: [
            "共橙一站式超市（南昌东祥路店）",
            "惠宜选超市（上饶信州大道店）"
        ],
        5: [
            "共橙一站式超市（南昌京东大道店）",
            "共橙一站式超市（德兴滨河大道店）",
            "惠宜选-江西上饶市婺源店"
        ]
    }

    # Column C is the 3rd column
    SHOP_COL_IDX = 3 

    print("Updating rows...")
    for row_idx, shop_list in updates.items():
        # Join with newline as per example in Row 2
        cell_value = "\n".join(shop_list)
        
        # Verify if the row index matches our expectation (Column A usually has ID)
        # Row 2 has ID 1. Row 3 should have ID 2.
        id_cell = ws.cell(row=row_idx, column=1)
        print(f"Row {row_idx}: ID column value is {id_cell.value}")
        
        target_cell = ws.cell(row=row_idx, column=SHOP_COL_IDX)
        print(f"  Filling Row {row_idx} Column {SHOP_COL_IDX} with {len(shop_list)} shops.")
        target_cell.value = cell_value

    try:
        wb.save(file_path)
        print("File saved successfully.")
    except PermissionError:
        print(f"Permission denied: {file_path} is likely open.")
        print("Please close the Excel file and run the script again.")
        # Attempt to save to a temp file to prove work is done
        backup_path = file_path.replace('.xlsx', '_updated.xlsx')
        try:
            wb.save(backup_path)
            print(f"Saved to backup file instead: {backup_path}")
        except Exception as e:
            print(f"Failed to save backup file: {e}")
    except Exception as e:
        print(f"Failed to save file: {e}")

if __name__ == '__main__':
    main()
