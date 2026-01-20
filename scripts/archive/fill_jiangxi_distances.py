# -*- coding: utf-8 -*-
"""
江西仓1月对账单距离填充脚本
根据可复用距离JSON文件，按照物流规则填充距离数据
输出格式：店名-距离km（与合肥仓格式一致）
"""

import pandas as pd
import json
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 配置路径
EXCEL_PATH = 'data/jiangxi/summary/2026/惠宜选江西仓1月份对账单0110.xlsx'
JSON_PATH = 'data/jiangxi/cache/reusable_distances.json'
OUTPUT_PATH = 'data/jiangxi/summary/2026/惠宜选江西仓1月份对账单0110.xlsx'

# 江西仓固定起点
START_POINT = "惠宜选南昌仓"


def normalize_store_name(name, aggressive=False):
    """标准化店名，处理常见的格式差异"""
    if pd.isna(name):
        return None
    name = str(name).strip()
    # 统一括号格式
    name = name.replace('（', '(').replace('）', ')')
    # 移除多余空格
    name = re.sub(r'\s+', '', name)

    if aggressive:
        # 更激进的标准化，用于模糊匹配
        # 处理"历臣"和"厉臣"的混用
        name = name.replace('历臣', '厉臣')
        # 处理连字符差异
        name = name.replace('-', '-').replace('—', '-').replace('－', '-')
        # 处理"共橙一站式"和"共橙-站式"的混用
        name = re.sub(r'共橙[-\s]*站式', '共橙一站式', name)
        # 处理"共橙超市"和"共橙一站式超市"
        name = name.replace('供橙超市', '共橙超市')
        name = name.replace('供橙一站式超市', '共橙一站式超市')

    return name


def find_distance(distances, from_store, to_store):
    """
    从距离字典中查找距离，尝试多种格式匹配
    返回 (距离值, 匹配的键) 或 (None, None)
    """
    # 标准化店名
    from_norm = normalize_store_name(from_store)
    to_norm = normalize_store_name(to_store)

    if not from_norm or not to_norm:
        return None, None

    # 构建可能的查询键变体
    key_variants = []

    # 原始格式
    key_variants.append(f"{from_store} -> {to_store}")
    key_variants.append(f"{from_norm} -> {to_norm}")

    # 尝试不同的括号格式
    from_half = from_store.replace('（', '(').replace('）', ')')
    to_half = to_store.replace('（', '(').replace('）', ')')
    key_variants.append(f"{from_half} -> {to_half}")

    # 尝试全角括号
    from_full = from_store.replace('(', '（').replace(')', '）')
    to_full = to_store.replace('(', '（').replace(')', '）')
    key_variants.append(f"{from_full} -> {to_full}")

    # 遍历所有变体查找
    for key in key_variants:
        if key in distances:
            return distances[key], key

    # 模糊匹配：使用激进标准化
    from_agg = normalize_store_name(from_store, aggressive=True)
    to_agg = normalize_store_name(to_store, aggressive=True)

    # 遍历所有距离记录进行模糊匹配
    for dist_key, dist_val in distances.items():
        if ' -> ' not in dist_key:
            continue
        parts = dist_key.split(' -> ')
        if len(parts) != 2:
            continue
        key_from, key_to = parts
        key_from_norm = normalize_store_name(key_from)
        key_to_norm = normalize_store_name(key_to)
        key_from_agg = normalize_store_name(key_from, aggressive=True)
        key_to_agg = normalize_store_name(key_to, aggressive=True)

        # 普通匹配
        if key_from_norm == from_norm and key_to_norm == to_norm:
            return dist_val, dist_key

        # 激进匹配
        if key_from_agg == from_agg and key_to_agg == to_agg:
            return dist_val, dist_key

    return None, None


def process_route(route_text, distances):
    """
    处理一条路线，返回每个站点的距离信息
    route_text: 包含换行符的站点列表
    返回: [(站点名, 距离, 查询键, 是否找到), ...]
    """
    if pd.isna(route_text):
        return []

    stops = [s.strip() for s in str(route_text).split('\n') if s.strip()]
    if not stops:
        return []

    results = []

    # 第一站：从起点到第一站
    first_stop = stops[0]
    dist, key = find_distance(distances, START_POINT, first_stop)
    results.append({
        'stop': first_stop,
        'from': START_POINT,
        'to': first_stop,
        'distance': dist,
        'key': key,
        'found': dist is not None
    })

    # 后续站点：前一站到下一站
    for i in range(1, len(stops)):
        prev_stop = stops[i-1]
        curr_stop = stops[i]
        dist, key = find_distance(distances, prev_stop, curr_stop)
        results.append({
            'stop': curr_stop,
            'from': prev_stop,
            'to': curr_stop,
            'distance': dist,
            'key': key,
            'found': dist is not None
        })

    return results


def main():
    print("=" * 60)
    print("江西仓1月对账单距离填充")
    print("=" * 60)

    # 读取距离数据
    print(f"\n读取距离数据: {JSON_PATH}")
    with open(JSON_PATH, 'r', encoding='utf-8') as f:
        distances = json.load(f)
    print(f"共加载 {len(distances)} 条距离记录")

    # 读取Excel数据
    print(f"\n读取Excel数据: {EXCEL_PATH}")
    df = pd.read_excel(EXCEL_PATH, header=None)
    print(f"Excel尺寸: {df.shape}")

    # 使用openpyxl处理以保留格式
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    # 统计信息
    total_routes = 0
    total_stops = 0
    found_count = 0
    not_found_count = 0
    not_found_details = []

    # 处理每一行（从第2行开始，跳过标题行）
    print("\n开始处理路线...")
    print("-" * 60)

    for row_idx in range(2, ws.max_row + 1):
        # 读取店名列（C列，索引3）
        store_cell = ws.cell(row=row_idx, column=3)
        route_text = store_cell.value

        if pd.isna(route_text) or not str(route_text).strip():
            continue

        # 检查是否已经包含距离信息（格式：店名-XXkm）
        # 如果第一行店名已经包含"-km"，说明该行已填充距离，跳过
        first_line = str(route_text).split('\n')[0].strip()
        if re.search(r'-\d+(\.\d+)?km', first_line) or re.search(r'-\?km', first_line):
            print(f"\n第{row_idx}行: 已包含距离信息，跳过")
            continue

        total_routes += 1
        results = process_route(route_text, distances)

        if not results:
            continue

        # 构建带距离的店名字符串
        formatted_stops = []

        print(f"\n第{row_idx}行 (第{total_routes}车):")
        for r in results:
            total_stops += 1
            if r['found']:
                found_count += 1
                # 格式化距离，保留合理的小数位
                dist_val = r['distance']
                if dist_val == int(dist_val):
                    dist_str = f"{int(dist_val)}km"
                else:
                    dist_str = f"{dist_val}km"
                formatted_stops.append(f"{r['stop']}-{dist_str}")
                print(f"  {r['stop']}-{dist_str} [找到]")
            else:
                not_found_count += 1
                formatted_stops.append(f"{r['stop']}-?km")
                not_found_details.append({
                    'row': row_idx,
                    'from': r['from'],
                    'to': r['to']
                })
                print(f"  {r['stop']}-?km [未找到]")

        # 将格式化后的店名（带距离）写回C列
        new_route_text = '\n'.join(formatted_stops)
        ws.cell(row=row_idx, column=3, value=new_route_text)

    # 保存结果
    print("\n" + "=" * 60)
    print("处理完成!")
    print(f"  总路线数: {total_routes}")
    print(f"  总站点数: {total_stops}")
    print(f"  找到距离: {found_count} ({found_count/total_stops*100:.1f}%)")
    print(f"  未找到距离: {not_found_count} ({not_found_count/total_stops*100:.1f}%)")

    # 保存文件
    print(f"\n保存结果到: {OUTPUT_PATH}")
    wb.save(OUTPUT_PATH)
    print("保存成功!")

    # 输出未找到的距离详情（便于后续补充）
    if not_found_details:
        print("\n未找到的距离段:")
        for item in not_found_details[:20]:  # 只显示前20条
            print(f"  行{item['row']}: {item['from']} -> {item['to']}")
        if len(not_found_details) > 20:
            print(f"  ... 还有 {len(not_found_details) - 20} 条未显示")


if __name__ == '__main__':
    main()
