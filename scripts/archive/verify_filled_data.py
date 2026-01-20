#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
验证填充后的数据
"""

from openpyxl import load_workbook

def verify_filled_data(excel_file):
    """验证填充后的数据"""

    print("="*100)
    print("验证填充后的对账单数据")
    print("="*100)

    wb = load_workbook(excel_file)
    ws = wb.active

    print("\n前10车的数据预览:\n")

    for row_idx in range(2, min(12, ws.max_row + 1)):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        print(f"车次 {vehicle_no}:")
        print("-"*80)

        if shop_names_cell:
            shop_lines = str(shop_names_cell).split('\n')
            for i, line in enumerate(shop_lines, 1):
                # 检查是否有距离
                if '-?km' in line:
                    status = "✗ 待填充"
                elif 'km' in line:
                    status = "✓ 已填充"
                else:
                    status = "？ 无距离"

                print(f"  {i}. {line} {status}")

        print()

    # 统计待填充的数量
    print("="*100)
    print("统计信息")
    print("="*100)

    total_segments = 0
    filled_segments = 0
    missing_segments = 0

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        if shop_names_cell:
            shop_lines = str(shop_names_cell).split('\n')
            for line in shop_lines:
                if line.strip():
                    total_segments += 1
                    if '-?km' in line:
                        missing_segments += 1
                    elif 'km' in line:
                        filled_segments += 1

    print(f"\n总路段数: {total_segments}")
    print(f"已填充: {filled_segments} ({filled_segments/total_segments*100:.1f}%)")
    print(f"待填充: {missing_segments} ({missing_segments/total_segments*100:.1f}%)")

    wb.close()


if __name__ == '__main__':
    excel_file = r'D:\Work\logistics\惠宜选合肥仓1月份对账单_已填充距离.xlsx'
    verify_filled_data(excel_file)
