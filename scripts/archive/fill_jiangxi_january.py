"""
填充江西1月份对账单数据
从txt文件读取物流店名数据,填充到Excel表格中
"""
import openpyxl
from datetime import datetime, timedelta

def parse_txt_data(txt_file):
    """
    解析txt文件,返回按日期分组的店名数据
    格式: {日期: [车辆列表]}
    """
    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    data = {}
    current_date = None
    current_vehicle = []

    for line in lines:
        line = line.strip()

        # 跳过空行
        if not line:
            # 如果当前有车辆数据,保存它
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []
            continue

        # 检查是否是日期行(格式如"1.2")
        if line.replace('.', '').isdigit() and '.' in line:
            # 保存前一辆车的数据
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []

            current_date = line
        else:
            # 店名行
            current_vehicle.append(line)

    # 保存最后一辆车的数据
    if current_vehicle and current_date:
        if current_date not in data:
            data[current_date] = []
        data[current_date].append(current_vehicle)

    return data

def date_str_to_excel_serial(date_str):
    """
    将日期字符串(如"1.2")转换为Excel日期序列号
    假设是2026年1月
    """
    parts = date_str.split('.')
    month = int(parts[0])
    day = int(parts[1])

    # 2026年1月
    date_obj = datetime(2026, month, day)

    # Excel的日期序列号从1899-12-30开始
    excel_epoch = datetime(1899, 12, 30)
    delta = date_obj - excel_epoch

    return delta.days

def fill_excel_data(excel_file, txt_data):
    """
    将txt数据填充到Excel文件中
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 找到第一个空白行
    first_empty_row = None
    for i in range(2, 500):
        if ws[f'B{i}'].value is None and ws[f'C{i}'].value is None:
            first_empty_row = i
            break

    if first_empty_row is None:
        print("未找到空白行!")
        return

    print(f"从第 {first_empty_row} 行开始填充数据")

    # 获取当前最大序号
    current_seq = ws[f'A{first_empty_row - 1}'].value
    if current_seq is None:
        # 向上查找最后一个有序号的行
        for i in range(first_empty_row - 1, 1, -1):
            if ws[f'A{i}'].value is not None:
                current_seq = ws[f'A{i}'].value
                break
        if current_seq is None:
            current_seq = 0

    current_row = first_empty_row

    # 按日期排序
    sorted_dates = sorted(txt_data.keys(), key=lambda x: tuple(map(int, x.split('.'))))

    for date_str in sorted_dates:
        vehicles = txt_data[date_str]
        excel_date = date_str_to_excel_serial(date_str)

        print(f"\n填充日期: {date_str} (Excel序列号: {excel_date})")
        print(f"  共 {len(vehicles)} 辆车")

        for vehicle_stores in vehicles:
            current_seq += 1

            # 填充序号
            ws[f'A{current_row}'] = current_seq

            # 填充日期
            ws[f'B{current_row}'] = excel_date

            # 填充店名(用换行符连接)
            stores_text = '\n'.join(vehicle_stores)
            ws[f'C{current_row}'] = stores_text

            print(f"  第{current_row}行: 序号={current_seq}, 店铺数={len(vehicle_stores)}, 首店={vehicle_stores[0][:20]}...")

            current_row += 1

    # 保存文件
    wb.save(excel_file)
    print(f"\n数据填充完成! 共填充 {current_row - first_empty_row} 行数据")
    print(f"保存到: {excel_file}")

if __name__ == '__main__':
    txt_file = 'data/jiangxi/details/物流店名数据_1.2-1.9.txt'
    excel_file = 'data/jiangxi/summary/惠宜选江西仓1月份对账单.xlsx'

    print("开始处理数据...")
    print(f"读取文件: {txt_file}")

    # 解析txt数据
    txt_data = parse_txt_data(txt_file)

    print(f"\n解析到 {len(txt_data)} 个日期的数据:")
    for date_str, vehicles in sorted(txt_data.items(), key=lambda x: tuple(map(int, x[0].split('.')))):
        print(f"  {date_str}: {len(vehicles)} 辆车")

    # 填充Excel
    fill_excel_data(excel_file, txt_data)
