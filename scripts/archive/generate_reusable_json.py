#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成可复用距离的JSON文件和报告
"""

import pandas as pd
from openpyxl import load_workbook
import json
import re
from collections import defaultdict

STARTING_POINT = "丰树合肥现代综合产业园"

def parse_shop_and_distance(shop_line):
    """解析店名行，提取店名和距离"""
    shop_line = shop_line.strip()
    pattern = r'^(.+?)-(\d+(?:\.\d+)?)km$'
    match = re.match(pattern, shop_line)

    if match:
        shop_name = match.group(1)
        distance = float(match.group(2))
        return shop_name, distance
    else:
        return shop_line, None


def extract_routes_from_excel(file_path, has_distance_in_name=False):
    """从对账单Excel中提取路线数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    routes = []

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        shops = []
        distances = []

        if shop_names_cell:
            shop_lines = [s.strip() for s in str(shop_names_cell).split('\n') if s.strip()]

            for shop_line in shop_lines:
                if has_distance_in_name:
                    shop_name, distance = parse_shop_and_distance(shop_line)
                    shops.append(shop_name)
                    distances.append(distance)
                else:
                    shops.append(shop_line)
                    distances.append(None)

        route = {
            'vehicle_no': vehicle_no,
            'date': date_value,
            'shops': shops,
            'distances': distances
        }

        routes.append(route)

    return routes


def build_segments(route):
    """根据README规则构建路段"""
    segments = []
    shops = route['shops']
    distances = route['distances']

    if not shops:
        return segments

    first_distance = distances[0] if len(distances) > 0 else None
    segments.append((STARTING_POINT, shops[0], first_distance))

    for i in range(1, len(shops)):
        distance = distances[i] if i < len(distances) else None
        segments.append((shops[i-1], shops[i], distance))

    return segments


print("生成可复用距离数据...")

# 读取12月份数据
dec_file = r'D:\Work\logistics\惠宜选合肥仓12月份对账单.xlsx'
dec_routes = extract_routes_from_excel(dec_file, has_distance_in_name=True)

# 构建距离字典
dec_distance_map = {}

for route in dec_routes:
    segments = build_segments(route)
    for seg in segments:
        key = (seg[0], seg[1])
        distance = seg[2]

        if distance is not None:
            if key not in dec_distance_map:
                dec_distance_map[key] = []
            dec_distance_map[key].append(distance)

# 生成JSON（使用字符串键）
reusable_json = {}
for key, distances in dec_distance_map.items():
    segment_key = f"{key[0]} -> {key[1]}"
    avg_distance = sum(distances) / len(distances)
    reusable_json[segment_key] = round(avg_distance, 2)

# 保存JSON
output_file = r'D:\Work\logistics\reusable_distances.json'
with open(output_file, 'w', encoding='utf-8') as f:
    json.dump(reusable_json, f, ensure_ascii=False, indent=2)

print(f"✓ JSON文件已保存: {output_file}")
print(f"✓ 包含 {len(reusable_json)} 个路段的距离数据")

# 生成统计报告
print("\n" + "="*100)
print("可复用距离统计报告")
print("="*100)

# 按起点分组统计
from_point_stats = defaultdict(list)
for segment_key, distance in reusable_json.items():
    parts = segment_key.split(" -> ")
    if len(parts) == 2:
        from_point = parts[0]
        from_point_stats[from_point].append((segment_key, distance))

print(f"\n从 '{STARTING_POINT}' 出发的路线: {len(from_point_stats.get(STARTING_POINT, []))} 条")
if STARTING_POINT in from_point_stats:
    for seg, dist in sorted(from_point_stats[STARTING_POINT], key=lambda x: x[1]):
        print(f"  {dist:6.2f} km -> {seg.split(' -> ')[1]}")

print(f"\n总计: {len(reusable_json)} 个路段可以从12月份复用")
print("\n数据已准备好，可用于填充1月份对账单的距离列")
