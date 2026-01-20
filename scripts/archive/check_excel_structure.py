#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
检查对账单Excel的结构
"""

from openpyxl import load_workbook
import pandas as pd

def check_excel_structure(file_path, file_label):
    """检查Excel结构"""
    print("="*100)
    print(f"检查 {file_label}")
    print("="*100)

    wb = load_workbook(file_path)
    ws = wb.active

    # 打印前10行的数据
    print("\n前10行数据:")
    print("-"*100)
    for row_idx in range(1, min(11, ws.max_row + 1)):
        print(f"\n第{row_idx}行:")
        for col_idx in range(1, min(15, ws.max_column + 1)):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                # 截断长字符串
                value_str = str(cell_value)
                if len(value_str) > 50:
                    value_str = value_str[:50] + "..."
                print(f"  列{col_idx}: {value_str}")

    # 统计有数据的列
    print("\n" + "="*100)
    print("列数据统计:")
    print("-"*100)
    for col_idx in range(1, min(15, ws.max_column + 1)):
        non_empty_count = 0
        for row_idx in range(2, min(50, ws.max_row + 1)):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                non_empty_count += 1

        print(f"列{col_idx}: 前50行中有 {non_empty_count} 个非空单元格")

    wb.close()


if __name__ == '__main__':
    jan_file = r'D:\Work\logistics\惠宜选合肥仓1月份对账单.xlsx'
    dec_file = r'D:\Work\logistics\惠宜选合肥仓12月份对账单.xlsx'

    check_excel_structure(jan_file, "1月份对账单")
    print("\n\n")
    check_excel_structure(dec_file, "12月份对账单")
