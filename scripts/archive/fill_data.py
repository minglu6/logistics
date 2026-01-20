import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook

# 读取源文件
source_file = r'D:\Work\logistics\临努1.1.xlsx'
df_source = pd.read_excel(source_file, header=None)

# 识别6车数据（通过空行分隔）
vehicles = []
current_vehicle = []

for idx, row in df_source.iterrows():
    if idx == 0:  # 跳过标题行
        continue

    if pd.isna(row[0]):  # 空行
        if current_vehicle:  # 如果当前车有数据
            vehicles.append(current_vehicle)
            current_vehicle = []
    else:  # 有数据的行
        current_vehicle.append({
            '序号': row[0],
            '城市': row[1],
            '编号': row[2],
            '店名': row[3],
            '金额': row[4]
        })

# 添加最后一车（如果有）
if current_vehicle:
    vehicles.append(current_vehicle)

print(f"识别到 {len(vehicles)} 车数据：")
for i, vehicle in enumerate(vehicles, 1):
    print(f"第{i}车：{len(vehicle)}个店")
    for store in vehicle:
        print(f"  {store['店名']}")
    print()

# 读取目标文件
target_file = r'D:\Work\logistics\惠宜选合肥仓1月份对账单.xlsx'
wb = load_workbook(target_file)
ws = wb.active

# 将日期转换为Excel日期序列号
# 2026年1月1日的Excel序列号是46023
date_value = 46023  # 1月1日

# 从第2行开始填充（第1行是表头，第2行开始是数据）
# 为每车数据创建一行
start_row = 2  # Excel从1开始，但openpyxl也从1开始，所以第2行就是数据的第一行

for i, vehicle in enumerate(vehicles, 1):
    row_idx = start_row + i - 1

    # 列A：序号
    ws.cell(row=row_idx, column=1, value=i)

    # 列B：日期
    ws.cell(row=row_idx, column=2, value=date_value)

    # 列C：店名（用换行符连接）
    shop_names = '\n'.join([store['店名'] for store in vehicle])
    ws.cell(row=row_idx, column=3, value=shop_names)

    print(f"填充第{i}车数据到第{row_idx}行")

# 保存文件
wb.save(target_file)
print(f"\n数据填充完成！共填充了{len(vehicles)}车数据")
