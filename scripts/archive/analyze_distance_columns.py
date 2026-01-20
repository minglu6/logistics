"""
详细分析Excel表格中公里数列的结构
"""
import openpyxl

wb = openpyxl.load_workbook('data/jiangxi/summary/惠宜选江西仓1月份对账单.xlsx')
ws = wb.active

print('详细查看前10行的公里数列:')
print('=' * 100)
for i in range(2, 12):
    stores = ws[f'C{i}'].value
    col_d = ws[f'D{i}'].value
    col_e = ws[f'E{i}'].value

    if stores:
        store_list = stores.split('\n')
        store_count = len(store_list)
        print(f'\n第{i}行: {store_count}个店铺')
        print(f'  店铺: {store_list[0][:30]}...')
        print(f'  D列: {col_d}')
        print(f'  E列: {col_e}')
    else:
        print(f'\n第{i}行: 空行')
        print(f'  D列: {col_d}')
        print(f'  E列: {col_e}')
