import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os

def extract_vehicles_from_file(file_path):
    """从Excel文件中提取车次数据"""
    df_source = pd.read_excel(file_path, header=None)

    vehicles = []
    current_vehicle = []

    for idx, row in df_source.iterrows():
        if idx == 0:  # 跳过标题行
            continue

        if pd.isna(row[0]):  # 空行
            if current_vehicle:  # 如果当前车有数据
                vehicles.append(current_vehicle)
                current_vehicle = []
        else:  # 有数据的行
            current_vehicle.append({
                '序号': row[0],
                '城市': row[1],
                '编号': row[2],
                '店名': row[3],
                '金额': row[4]
            })

    # 添加最后一车（如果有）
    if current_vehicle:
        vehicles.append(current_vehicle)

    return vehicles

# 定义文件列表和对应的日期
files_to_process = [
    {'file': r'D:\Work\logistics\临努1.2xlsx.xlsx', 'day': 2, 'date_serial': 46024},
    {'file': r'D:\Work\logistics\临努1.3.xlsx', 'day': 3, 'date_serial': 46025},
    {'file': r'D:\Work\logistics\临努1.4.xlsx', 'day': 4, 'date_serial': 46026},
    {'file': r'D:\Work\logistics\临努1.5.xlsx', 'day': 5, 'date_serial': 46027},
    {'file': r'D:\Work\logistics\临努1.6.xlsx', 'day': 6, 'date_serial': 46028},
    {'file': r'D:\Work\logistics\临努1.8.xlsx', 'day': 8, 'date_serial': 46030},
]

# 读取目标文件
target_file = r'D:\Work\logistics\惠宜选合肥仓1月份对账单.xlsx'
wb = load_workbook(target_file)
ws = wb.active

# 当前起始行（1月1日已经填充了6车，占用了2-7行，所以从第8行开始）
current_row = 8
total_vehicles_added = 0

print("开始批量处理数据...\n")
print("="*80)

for file_info in files_to_process:
    file_path = file_info['file']
    day = file_info['day']
    date_serial = file_info['date_serial']

    if not os.path.exists(file_path):
        print(f"警告: 文件不存在 - {file_path}")
        continue

    print(f"\n处理: 1月{day}日 - {os.path.basename(file_path)}")

    # 提取车次数据
    vehicles = extract_vehicles_from_file(file_path)
    print(f"  识别到 {len(vehicles)} 车数据")

    # 填充每一车的数据
    for i, vehicle in enumerate(vehicles, 1):
        # 列A：序号
        ws.cell(row=current_row, column=1, value=current_row - 1)

        # 列B：日期
        ws.cell(row=current_row, column=2, value=date_serial)

        # 列C：店名（用换行符连接）
        shop_names = '\n'.join([store['店名'] for store in vehicle])
        ws.cell(row=current_row, column=3, value=shop_names)

        print(f"    第{i}车：{len(vehicle)}个店 -> 填充到第{current_row}行")

        current_row += 1
        total_vehicles_added += 1

# 保存文件
wb.save(target_file)

print("\n" + "="*80)
print(f"\n批量处理完成！")
print(f"  总共处理了 {len(files_to_process)} 个文件")
print(f"  添加了 {total_vehicles_added} 车数据")
print(f"  数据填充到第 {current_row - 1} 行")
