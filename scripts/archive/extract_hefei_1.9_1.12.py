#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从临努1.9-1.12的Excel文件中提取物流店名数据
输出格式：按日期分组，每车用空行分隔
"""
import openpyxl
import os

def extract_stores_from_excel(excel_file):
    """
    从Excel文件中提取店名，按车辆分组
    返回：[[车1的店名列表], [车2的店名列表], ...]
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    vehicles = []
    current_vehicle = []

    # 从第2行开始读取（跳过标题行）
    for row_idx in range(2, ws.max_row + 1):
        # 读取店名（第4列）
        store_name = ws.cell(row=row_idx, column=4).value

        # 如果是空行，表示一车结束
        if store_name is None or str(store_name).strip() == '':
            if current_vehicle:
                vehicles.append(current_vehicle)
                current_vehicle = []
        else:
            current_vehicle.append(str(store_name).strip())

    # 保存最后一车
    if current_vehicle:
        vehicles.append(current_vehicle)

    return vehicles


def main():
    # 配置文件路径
    base_dir = 'data/hefei/details'
    output_file = 'data/hefei/details/物流店名数据_1.9-1.12.txt'

    # 要处理的文件列表（按日期顺序）
    files = [
        ('1.9', os.path.join(base_dir, '临努1.9.xlsx')),
        ('1.10', os.path.join(base_dir, '临努1.10.xlsx')),
        ('1.11', os.path.join(base_dir, '临努1.11.xlsx')),
        ('1.12', os.path.join(base_dir, '临努1.12.xlsx')),
    ]

    print("=" * 60)
    print("提取合肥临努1.9-1.12物流店名数据")
    print("=" * 60)

    all_data = {}

    for date_str, file_path in files:
        if not os.path.exists(file_path):
            print(f"\n警告: 文件不存在 - {file_path}")
            continue

        print(f"\n处理 {date_str}: {file_path}")
        vehicles = extract_stores_from_excel(file_path)
        all_data[date_str] = vehicles
        print(f"  提取到 {len(vehicles)} 辆车")
        for i, vehicle in enumerate(vehicles, 1):
            print(f"    第{i}车: {len(vehicle)}个店铺, 首店: {vehicle[0][:30]}...")

    # 写入输出文件
    print(f"\n写入输出文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        for date_str in ['1.9', '1.10', '1.11', '1.12']:
            if date_str not in all_data:
                continue

            # 写入日期标题
            f.write(f"{date_str}\n")
            f.write("\n")

            # 写入该日期的所有车辆
            vehicles = all_data[date_str]
            for vehicle_stores in vehicles:
                for store in vehicle_stores:
                    f.write(f"{store}\n")
                f.write("\n")  # 车辆之间用空行分隔

    print("=" * 60)
    print("提取完成!")
    print(f"输出文件: {output_file}")
    print("=" * 60)


if __name__ == '__main__':
    main()
