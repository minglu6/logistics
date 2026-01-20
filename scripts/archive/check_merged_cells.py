# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook('data/jiangxi/summary/惠宜选江西仓1月份对账单.xlsx')
ws = wb.active

print('检查合并单元格信息:')
print(f'合并单元格总数: {len(ws.merged_cells.ranges)}')

# 找出C列（店名列）的合并单元格
store_merges = []
for merge_range in ws.merged_cells.ranges:
    if merge_range.min_col == 3:  # C列
        store_merges.append({
            'start_row': merge_range.min_row,
            'end_row': merge_range.max_row,
            'value': ws.cell(row=merge_range.min_row, column=3).value
        })

print(f'\nC列合并单元格数量: {len(store_merges)}')
print('\n前10个合并单元格:')
for i, m in enumerate(sorted(store_merges, key=lambda x: x['start_row'])[:10]):
    val = m['value'] if m['value'] else '(空)'
    val_preview = str(val).replace('\n', ' | ')[:60] + '...' if len(str(val)) > 60 else str(val).replace('\n', ' | ')
    print(f"  行{m['start_row']}-{m['end_row']}: {val_preview}")
