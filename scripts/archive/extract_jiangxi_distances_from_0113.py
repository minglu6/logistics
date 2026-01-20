#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从江西1月份对账单0113中提取距离数据，并合并到现有缓存中
按照物流规则提取"上一站 -> 下一站"的距离
如果有冲突且距离相差5km内，以新数据为准
"""

import pandas as pd
from openpyxl import load_workbook
import json
import re
from collections import defaultdict

STARTING_POINT = "惠宜选南昌仓"
CONFLICT_THRESHOLD = 5.0  # 5km内视为相同路段

def parse_shop_and_distance(shop_line):
    """解析店名行，提取店名和距离"""
    shop_line = shop_line.strip()
    # 匹配格式：店名-距离km
    pattern = r'^(.+?)-(\d+(?:\.\d+)?)km$'
    match = re.match(pattern, shop_line)

    if match:
        shop_name = match.group(1)
        distance = float(match.group(2))
        return shop_name, distance
    else:
        return shop_line, None


def extract_routes_from_excel(file_path):
    """从对账单Excel中提取路线数据"""
    wb = load_workbook(file_path)
    ws = wb.active
    routes = []

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        shops = []
        distances = []

        if shop_names_cell:
            shop_lines = [s.strip() for s in str(shop_names_cell).split('\n') if s.strip()]

            for shop_line in shop_lines:
                shop_name, distance = parse_shop_and_distance(shop_line)
                shops.append(shop_name)
                distances.append(distance)

        route = {
            'vehicle_no': vehicle_no,
            'date': date_value,
            'shops': shops,
            'distances': distances
        }

        routes.append(route)

    return routes


def build_segments(route):
    """根据README规则构建路段"""
    segments = []
    shops = route['shops']
    distances = route['distances']

    if not shops:
        return segments

    # 第一段：起点到第一站
    first_distance = distances[0] if len(distances) > 0 else None
    segments.append((STARTING_POINT, shops[0], first_distance))

    # 后续段：前一站到下一站
    for i in range(1, len(shops)):
        distance = distances[i] if i < len(distances) else None
        segments.append((shops[i-1], shops[i], distance))

    return segments


print("=" * 80)
print("从江西对账单0113提取距离数据并更新缓存")
print("=" * 80)

# 文件路径
excel_file = 'data/jiangxi/summary/2026/惠宜选江西仓1月份对账单0113.xlsx'
cache_file = 'data/jiangxi/cache/reusable_distances.json'
large_diff_file = 'data/jiangxi/cache/large_distance_differences.json'

# 读取现有缓存
print(f"\n读取现有缓存: {cache_file}")
try:
    with open(cache_file, 'r', encoding='utf-8') as f:
        existing_distances = json.load(f)
    print(f"  现有缓存包含 {len(existing_distances)} 个路段")
except FileNotFoundError:
    print("  缓存文件不存在，将创建新文件")
    existing_distances = {}

# 提取新数据
print(f"\n处理Excel文件: {excel_file}")
routes = extract_routes_from_excel(excel_file)
print(f"  提取到 {len(routes)} 条路线")

# 构建新数据的距离字典（每个路段取平均值）
new_distance_map = defaultdict(list)
for route in routes:
    segments = build_segments(route)
    for seg in segments:
        key = (seg[0], seg[1])
        distance = seg[2]
        if distance is not None:
            new_distance_map[key].append(distance)

# 计算平均距离
new_avg_distances = {}
for key, distances in new_distance_map.items():
    avg_distance = sum(distances) / len(distances)
    new_avg_distances[key] = round(avg_distance, 2)

print(f"  新数据包含 {len(new_avg_distances)} 个路段")

# 合并数据
conflicts = []
large_differences = []
new_additions = 0
updates = 0

for key, new_distance in new_avg_distances.items():
    segment_key = f"{key[0]} -> {key[1]}"

    if segment_key in existing_distances:
        old_distance = existing_distances[segment_key]
        diff = abs(new_distance - old_distance)

        if diff <= CONFLICT_THRESHOLD:
            # 距离相差在5km内，使用新数据
            if diff > 0.1:  # 差异大于0.1km才记录
                conflicts.append({
                    'segment': segment_key,
                    'old_distance': old_distance,
                    'new_distance': new_distance,
                    'difference': diff
                })
            existing_distances[segment_key] = new_distance
            updates += 1
        else:
            # 距离相差超过5km，记录但保留旧数据
            large_differences.append({
                'segment': segment_key,
                'old_distance': old_distance,
                'new_distance': new_distance,
                'difference': diff,
                'used_distance': old_distance
            })
    else:
        # 新路段
        existing_distances[segment_key] = new_distance
        new_additions += 1

# 保存更新后的缓存
with open(cache_file, 'w', encoding='utf-8') as f:
    json.dump(existing_distances, f, ensure_ascii=False, indent=2)

print(f"\n✓ 缓存已更新: {cache_file}")
print(f"  总路段数: {len(existing_distances)}")
print(f"  新增路段: {new_additions}")
print(f"  更新路段: {updates}")

# 保存大差异数据
if large_differences:
    # 读取现有的大差异数据
    existing_large_diffs = []
    try:
        with open(large_diff_file, 'r', encoding='utf-8') as f:
            existing_large_diffs = json.load(f)
    except FileNotFoundError:
        pass

    # 添加新的大差异
    for d in large_differences:
        existing_large_diffs.append({
            'segment': d['segment'],
            'old_distance_km': d['old_distance'],
            'new_distance_km': d['new_distance'],
            'difference_km': round(d['difference'], 2),
            'used_distance_km': d['used_distance'],
            'source': '对账单0113',
            'warning': '距离差异超过5km，已保留旧数据，请人工检查'
        })

    with open(large_diff_file, 'w', encoding='utf-8') as f:
        json.dump(existing_large_diffs, f, ensure_ascii=False, indent=2)

    print(f"\n✓ 大差异数据已保存: {large_diff_file}")
    print(f"  包含 {len(existing_large_diffs)} 个需要检查的路段")

# 冲突报告
if conflicts:
    print("\n" + "=" * 80)
    print(f"发现 {len(conflicts)} 个冲突路段（距离相差5km内，已使用新数据）")
    print("=" * 80)
    for c in conflicts[:10]:
        print(f"\n路段: {c['segment']}")
        print(f"  旧数据: {c['old_distance']} km")
        print(f"  新数据: {c['new_distance']} km (已更新)")
        print(f"  差值: {c['difference']:.2f} km")
    if len(conflicts) > 10:
        print(f"\n... 还有 {len(conflicts) - 10} 个冲突路段未显示")

# 大差异报告
if large_differences:
    print("\n" + "=" * 80)
    print(f"警告：发现 {len(large_differences)} 个路段距离相差超过5km，已保留旧数据")
    print("=" * 80)
    for d in large_differences[:10]:
        print(f"\n路段: {d['segment']}")
        print(f"  旧数据: {d['old_distance']} km (保留)")
        print(f"  新数据: {d['new_distance']} km")
        print(f"  差值: {d['difference']:.2f} km ⚠️")
    if len(large_differences) > 10:
        print(f"\n... 还有 {len(large_differences) - 10} 个大差异路段未显示")
    print("\n这些路段已保留旧数据，请人工检查后决定使用哪个数据！")

print("\n" + "=" * 80)
print("提取完成！")
print("=" * 80)
