#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将可复用的距离数据填充到1月份对账单中
按照12月份的格式：店名-距离km
"""

import argparse
import json
import os
import re
from openpyxl import load_workbook
from collections import defaultdict

# 固定起点
STARTING_POINT = "丰树合肥现代综合产业园"

def load_reusable_distances(json_file):
    """加载可复用距离数据"""
    with open(json_file, 'r', encoding='utf-8') as f:
        distances = json.load(f)
    print(f"✓ 加载了 {len(distances)} 个路段的距离数据")
    return distances


def parse_shop_name(shop_line):
    """解析店名行，去除可能存在的距离后缀"""
    shop_line = shop_line.strip()
    if not shop_line:
        return shop_line

    pattern_distance = r'^(.+?)-(\d+(?:\.\d+)?)km$'
    pattern_missing = r'^(.+?)-\?km$'

    match = re.match(pattern_distance, shop_line)
    if match:
        return match.group(1)

    match = re.match(pattern_missing, shop_line)
    if match:
        return match.group(1)

    return shop_line


def build_output_path(excel_file):
    """生成输出文件路径"""
    base_name, ext = os.path.splitext(excel_file)
    return f"{base_name}_已填充距离{ext}"


def fill_distances_to_excel(excel_file, distance_map, starting_point, output_file=None):
    """将距离填充到Excel的店名列"""

    print(f"\n开始处理Excel文件: {excel_file}")

    # 加载Excel
    wb = load_workbook(excel_file)
    ws = wb.active

    # 统计
    total_routes = 0
    total_segments = 0
    filled_segments = 0
    missing_segments = 0
    missing_details = []

    # 从第2行开始处理（第1行是表头）
    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        # 如果序号为空，说明数据结束
        if vehicle_no is None:
            break

        total_routes += 1

        # 解析店名（用换行符分隔）
        if not shop_names_cell:
            continue

        shop_lines = [parse_shop_name(s) for s in str(shop_names_cell).split('\n') if s.strip()]

        if not shop_lines:
            continue

        # 构建新的店名列表（带距离）
        new_shop_lines = []

        # 第一段：起点 -> 第一个店
        first_shop = shop_lines[0]
        segment_key = f"{starting_point} -> {first_shop}"
        total_segments += 1

        if segment_key in distance_map:
            distance = distance_map[segment_key]
            new_shop_lines.append(f"{first_shop}-{distance}km")
            filled_segments += 1
        else:
            new_shop_lines.append(f"{first_shop}-?km")
            missing_segments += 1
            missing_details.append({
                'vehicle': vehicle_no,
                'segment': segment_key
            })

        # 后续段：前一站 -> 下一站
        for i in range(1, len(shop_lines)):
            prev_shop = shop_lines[i-1]
            curr_shop = shop_lines[i]
            segment_key = f"{prev_shop} -> {curr_shop}"
            total_segments += 1

            if segment_key in distance_map:
                distance = distance_map[segment_key]
                new_shop_lines.append(f"{curr_shop}-{distance}km")
                filled_segments += 1
            else:
                new_shop_lines.append(f"{curr_shop}-?km")
                missing_segments += 1
                missing_details.append({
                    'vehicle': vehicle_no,
                    'segment': segment_key
                })

        # 更新店名列（用换行符连接）
        new_shop_text = '\n'.join(new_shop_lines)
        ws.cell(row=row_idx, column=3, value=new_shop_text)

    # 保存文件
    if output_file is None:
        output_file = build_output_path(excel_file)
    wb.save(output_file)

    # 打印统计结果
    print("\n" + "="*100)
    print("填充完成统计")
    print("="*100)
    print(f"处理车次数: {total_routes}")
    print(f"总路段数: {total_segments}")
    print(f"已填充路段: {filled_segments} ({filled_segments/total_segments*100:.1f}%)")
    print(f"未填充路段: {missing_segments} ({missing_segments/total_segments*100:.1f}%)")
    print(f"\n文件已保存: {output_file}")

    # 输出缺失的路段详情
    if missing_details:
        print("\n" + "="*100)
        print("未找到距离的路段（需要手动填充）")
        print("="*100)

        # 按路段分组
        segment_groups = defaultdict(list)
        for detail in missing_details:
            segment_groups[detail['segment']].append(detail['vehicle'])

        print(f"\n共 {len(segment_groups)} 个不同的路段需要手动填充:\n")
        for i, (segment, vehicles) in enumerate(sorted(segment_groups.items()), 1):
            vehicle_str = ", ".join(map(str, sorted(set(vehicles))))
            print(f"{i}. {segment}")
            print(f"   出现在车次: {vehicle_str}")

    return output_file, filled_segments, missing_segments


def main():
    parser = argparse.ArgumentParser(description="将可复用距离填充到对账单店名列")
    parser.add_argument(
        "--json-file",
        default=r"D:\Work\logistics\reusable_distances.json",
        help="可复用距离JSON文件路径",
    )
    parser.add_argument(
        "--excel-file",
        default=r"D:\Work\logistics\惠宜选合肥仓1月份对账单.xlsx",
        help="需要填充的对账单Excel路径",
    )
    parser.add_argument(
        "--starting-point",
        default=STARTING_POINT,
        help="固定起点名称",
    )
    parser.add_argument(
        "--output-file",
        default=None,
        help="输出文件路径（默认自动生成）",
    )
    args = parser.parse_args()

    print("="*100)
    print("将距离数据填充到1月份对账单")
    print("="*100)

    if not os.path.exists(args.excel_file):
        print(f"Excel文件不存在: {args.excel_file}")
        return

    if not os.path.exists(args.json_file):
        print(f"JSON文件不存在: {args.json_file}")
        return

    # 加载可复用距离
    distance_map = load_reusable_distances(args.json_file)

    # 填充到Excel
    output_file, filled, missing = fill_distances_to_excel(
        args.excel_file,
        distance_map,
        args.starting_point,
        args.output_file,
    )

    print("\n" + "="*100)
    print("处理完成")
    print("="*100)
    print(f"\n✓ 新文件: {output_file}")
    print(f"✓ 已填充: {filled} 个路段")
    print(f"✗ 待填充: {missing} 个路段（标记为 ?km）")
    print(f"\n下一步: 打开新文件，手动填充标记为 '?km' 的路段距离")


if __name__ == '__main__':
    main()
