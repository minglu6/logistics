"""
验证江西1月份对账单数据填充结果
"""
import openpyxl
from datetime import datetime, timedelta

excel_file = 'data/jiangxi/summary/惠宜选江西仓1月份对账单.xlsx'

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print('验证填充的数据:')
print('=' * 80)

for i in range(7, 25):
    seq = ws[f'A{i}'].value
    date_val = ws[f'B{i}'].value
    stores = ws[f'C{i}'].value

    if date_val:
        date_obj = datetime(1899, 12, 30) + timedelta(days=date_val)
        date_str = date_obj.strftime('%Y-%m-%d')
    else:
        date_str = '无'

    if stores:
        stores_list = stores.split('\n')
        stores_preview = f'{stores_list[0][:30]}... (共{len(stores_list)}个店)'
    else:
        stores_preview = '无'

    print(f'第{i}行: 序号={seq}, 日期={date_str}, 店名={stores_preview}')

print('=' * 80)
print('数据填充验证完成!')
