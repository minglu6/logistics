#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
分析1月和12月对账单中可复用的距离数据
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
import json
from collections import defaultdict

# 固定起点
STARTING_POINT = "丰树合肥现代综合产业园"

def extract_routes_from_excel(file_path):
    """从对账单Excel中提取路线数据

    Returns:
        list: 路线列表，每个路线是一个字典：
        {
            'date': 日期,
            'vehicle_no': 车次序号,
            'shops': [店名列表],
            'distances': [距离列表] (如果有的话)
        }
    """
    wb = load_workbook(file_path)
    ws = wb.active

    routes = []

    # 从第2行开始读取（第1行是表头）
    for row_idx in range(2, ws.max_row + 1):
        # 读取序号、日期、店名、公里数列
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names = ws.cell(row=row_idx, column=3).value
        distance_col1 = ws.cell(row=row_idx, column=4).value  # 第一个公里数列
        distance_col2 = ws.cell(row=row_idx, column=5).value  # 第二个公里数列

        # 如果序号为空，说明数据结束
        if vehicle_no is None:
            break

        # 解析店名（用换行符分隔）
        if shop_names:
            shops = [s.strip() for s in str(shop_names).split('\n') if s.strip()]
        else:
            shops = []

        # 尝试解析距离数据
        distances = []
        if distance_col1:
            # 假设距离也是用换行符分隔的
            dist_str = str(distance_col1)
            for d in dist_str.split('\n'):
                d = d.strip()
                if d:
                    try:
                        distances.append(float(d))
                    except:
                        distances.append(None)

        route = {
            'vehicle_no': vehicle_no,
            'date': date_value,
            'shops': shops,
            'distances': distances if distances else [None] * len(shops)
        }

        routes.append(route)

    return routes


def build_segments(route):
    """根据README规则构建路段

    Args:
        route: 路线字典

    Returns:
        list: 路段列表，每个路段是 (起点, 终点, 距离)
    """
    segments = []
    shops = route['shops']
    distances = route['distances']

    if not shops:
        return segments

    # 第一段：起点 -> 第一个店
    first_distance = distances[0] if len(distances) > 0 else None
    segments.append((STARTING_POINT, shops[0], first_distance))

    # 后续段：前一站 -> 下一站
    for i in range(1, len(shops)):
        distance = distances[i] if i < len(distances) else None
        segments.append((shops[i-1], shops[i], distance))

    return segments


def analyze_reusable_distances(jan_file, dec_file):
    """分析可复用的距离"""

    print("="*100)
    print("分析1月和12月对账单中的可复用距离")
    print("="*100)

    # 读取1月份数据
    print("\n读取1月份对账单...")
    jan_routes = extract_routes_from_excel(jan_file)
    print(f"  找到 {len(jan_routes)} 条路线")

    # 读取12月份数据
    print("\n读取12月份对账单...")
    dec_routes = extract_routes_from_excel(dec_file)
    print(f"  找到 {len(dec_routes)} 条路线")

    # 构建1月份的路段
    print("\n构建1月份路段...")
    jan_segments = []
    for route in jan_routes:
        segments = build_segments(route)
        for seg in segments:
            jan_segments.append({
                'vehicle_no': route['vehicle_no'],
                'date': route['date'],
                'from': seg[0],
                'to': seg[1],
                'distance': seg[2]
            })
    print(f"  总共 {len(jan_segments)} 个路段")

    # 构建12月份的路段和距离字典
    print("\n构建12月份路段和距离字典...")
    dec_distance_map = {}  # {(起点, 终点): [距离列表]}
    dec_segments = []

    for route in dec_routes:
        segments = build_segments(route)
        for seg in segments:
            key = (seg[0], seg[1])
            distance = seg[2]

            dec_segments.append({
                'vehicle_no': route['vehicle_no'],
                'date': route['date'],
                'from': seg[0],
                'to': seg[1],
                'distance': distance
            })

            # 只记录有距离数据的路段
            if distance is not None:
                if key not in dec_distance_map:
                    dec_distance_map[key] = []
                dec_distance_map[key].append(distance)

    print(f"  总共 {len(dec_segments)} 个路段")
    print(f"  其中有距离数据的路段: {len(dec_distance_map)} 个")

    # 查找可复用的距离
    print("\n" + "="*100)
    print("分析可复用的距离")
    print("="*100)

    reusable_count = 0
    missing_count = 0
    reusable_segments = []
    missing_segments = []

    for jan_seg in jan_segments:
        key = (jan_seg['from'], jan_seg['to'])

        # 如果1月份这个路段没有距离，但12月份有
        if jan_seg['distance'] is None and key in dec_distance_map:
            dec_distances = dec_distance_map[key]
            avg_distance = sum(dec_distances) / len(dec_distances)

            reusable_segments.append({
                'jan_vehicle': jan_seg['vehicle_no'],
                'jan_date': jan_seg['date'],
                'from': jan_seg['from'],
                'to': jan_seg['to'],
                'dec_distance': avg_distance,
                'dec_occurrences': len(dec_distances)
            })
            reusable_count += 1

        # 如果1月份和12月份都没有距离
        elif jan_seg['distance'] is None and key not in dec_distance_map:
            missing_segments.append({
                'jan_vehicle': jan_seg['vehicle_no'],
                'jan_date': jan_seg['date'],
                'from': jan_seg['from'],
                'to': jan_seg['to']
            })
            missing_count += 1

    # 输出结果
    print(f"\n可从12月份复用的路段: {reusable_count} 个")
    print(f"仍需填充的路段: {missing_count} 个")

    if reusable_segments:
        print("\n" + "="*100)
        print("可复用的距离明细")
        print("="*100)

        # 按路段分组
        segment_groups = defaultdict(list)
        for seg in reusable_segments:
            key = f"{seg['from']} -> {seg['to']}"
            segment_groups[key].append(seg)

        for segment_key, occurrences in segment_groups.items():
            avg_dist = occurrences[0]['dec_distance']
            dec_count = occurrences[0]['dec_occurrences']
            jan_count = len(occurrences)

            print(f"\n路段: {segment_key}")
            print(f"  12月份平均距离: {avg_dist:.2f} km (出现 {dec_count} 次)")
            print(f"  1月份出现次数: {jan_count}")
            print(f"  1月份车次: ", end="")
            vehicles = [str(occ['jan_vehicle']) for occ in occurrences]
            print(", ".join(vehicles))

    if missing_segments:
        print("\n" + "="*100)
        print("仍需填充距离的路段（1月和12月都没有数据）")
        print("="*100)

        # 按路段分组
        segment_groups = defaultdict(list)
        for seg in missing_segments:
            key = f"{seg['from']} -> {seg['to']}"
            segment_groups[key].append(seg)

        print(f"\n总共 {len(segment_groups)} 个不同的路段需要填充")
        print("\n前20个需要填充的路段:")
        for i, (segment_key, occurrences) in enumerate(list(segment_groups.items())[:20], 1):
            jan_count = len(occurrences)
            print(f"{i}. {segment_key} (1月份出现 {jan_count} 次)")

    # 生成JSON格式的可复用距离数据
    print("\n" + "="*100)
    print("生成可复用距离的JSON数据")
    print("="*100)

    reusable_json = {}
    for segment_key, occurrences in segment_groups.items():
        if occurrences and 'dec_distance' in occurrences[0]:
            reusable_json[segment_key] = occurrences[0]['dec_distance']

    output_file = r'D:\Work\logistics\reusable_distances.json'
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(reusable_json, f, ensure_ascii=False, indent=2)

    print(f"\n可复用距离已保存到: {output_file}")
    print(f"包含 {len(reusable_json)} 个路段的距离数据")

    return reusable_segments, missing_segments


if __name__ == '__main__':
    jan_file = r'D:\Work\logistics\惠宜选合肥仓1月份对账单.xlsx'
    dec_file = r'D:\Work\logistics\惠宜选合肥仓12月份对账单.xlsx'

    reusable, missing = analyze_reusable_distances(jan_file, dec_file)
