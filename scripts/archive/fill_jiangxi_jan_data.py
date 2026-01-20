#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
将物流店名数据从txt文件填充到江西1月份对账单Excel中
"""

import openpyxl
import os
import re

def parse_txt_file(file_path):
    """解析txt文件，返回按日期和车辆组织的数据"""
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 按日期分割
    date_pattern = r'1\.(\d+)'
    dates_data = {}

    lines = content.strip().split('\n')
    current_date = None
    current_vehicle_shops = []
    all_vehicles = []

    for line in lines:
        line = line.strip()

        # 检查是否是日期行
        match = re.match(date_pattern, line)
        if match:
            # 保存前一个日期的数据
            if current_date and all_vehicles:
                dates_data[current_date] = all_vehicles

            # 开始新日期
            current_date = int(match.group(1))
            all_vehicles = []
            current_vehicle_shops = []
        elif line == '':
            # 空行表示一辆车的结束
            if current_vehicle_shops:
                all_vehicles.append(current_vehicle_shops)
                current_vehicle_shops = []
        elif current_date is not None:
            # 店名行
            current_vehicle_shops.append(line)

    # 保存最后一个日期的数据
    if current_vehicle_shops:
        all_vehicles.append(current_vehicle_shops)
    if current_date and all_vehicles:
        dates_data[current_date] = all_vehicles

    return dates_data


def fill_excel(excel_path, dates_data):
    """将数据填充到Excel文件中"""
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 1月1日对应的序列号是46023
    base_date_serial = 46023

    # 找到当前最大行号
    current_row = 2
    while ws.cell(row=current_row, column=1).value is not None:
        current_row += 1

    print(f"开始从第 {current_row} 行填充数据...\n")

    # 按日期顺序填充
    for day in sorted(dates_data.keys()):
        vehicles = dates_data[day]
        date_serial = base_date_serial + day - 1

        print(f"填充 1月{day}日 数据 (序列号: {date_serial})，共 {len(vehicles)} 辆车")

        for vehicle_idx, shops in enumerate(vehicles, start=1):
            if not shops:
                continue

            # 填充序号
            ws.cell(row=current_row, column=1).value = vehicle_idx

            # 填充日期（使用序列号）
            ws.cell(row=current_row, column=2).value = date_serial

            # 填充店名（用换行符连接）
            shop_names = '\n'.join(shops)
            ws.cell(row=current_row, column=3).value = shop_names

            print(f"  第 {vehicle_idx} 辆车: {len(shops)} 个店铺")

            current_row += 1

        print()

    # 保存文件
    try:
        wb.save(excel_path)
        print(f"✓ 数据已成功保存到: {excel_path}")
    except PermissionError:
        backup_path = excel_path.replace('.xlsx', '_filled.xlsx')
        wb.save(backup_path)
        print(f"原文件被占用，已保存到: {backup_path}")


def main():
    txt_file = r'D:\Work\logistics\data\jiangxi\details\物流店名数据_1.2-1.9.txt'
    excel_file = r'D:\Work\logistics\data\jiangxi\summary\惠宜选江西仓1月份对账单.xlsx'

    print("="*80)
    print("开始填充江西1月份对账单数据")
    print("="*80)

    # 解析txt文件
    print(f"\n1. 解析txt文件: {txt_file}")
    dates_data = parse_txt_file(txt_file)

    print(f"\n解析结果:")
    for day in sorted(dates_data.keys()):
        vehicles = dates_data[day]
        print(f"  1月{day}日: {len(vehicles)} 辆车")
        for i, shops in enumerate(vehicles, start=1):
            print(f"    第{i}辆车: {len(shops)} 个店铺")

    # 填充Excel
    print(f"\n2. 填充Excel文件: {excel_file}")
    fill_excel(excel_file, dates_data)

    print("\n" + "="*80)
    print("填充完成！")
    print("="*80)


if __name__ == '__main__':
    main()
