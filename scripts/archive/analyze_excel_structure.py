"""
分析Excel表格结构
"""
import openpyxl

wb = openpyxl.load_workbook('data/jiangxi/summary/惠宜选江西仓1月份对账单.xlsx')
ws = wb.active

# 查看第2行(1月1日第一车)的店名和公里数列
print('示例: 第2行数据结构')
print('店名列(C2):', ws['C2'].value[:100] if ws['C2'].value else None)
print('公里数列D(D2):', ws['D2'].value)
print('公里数列E(E2):', ws['E2'].value)

print('\n第6行(1月2日第一车):')
print('店名列(C6):', ws['C6'].value[:100] if ws['C6'].value else None)
print('公里数列D(D6):', ws['D6'].value)
print('公里数列E(E6):', ws['E6'].value)
