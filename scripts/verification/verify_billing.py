#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
对账单数据验证工具
支持多种验证模式：填充验证、数据统计、完整性检查等

用法:
    python -m scripts.verification.verify_billing --mode filled --input 对账单.xlsx
    python -m scripts.verification.verify_billing --mode summary --input 对账单.xlsx
    python -m scripts.verification.verify_billing --mode complete --input 对账单.xlsx
"""

import argparse
import os
import re
import sys
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.utils.common import parse_shop_and_distance


def verify_filled_data(excel_file, preview_rows=10):
    """
    验证填充后的数据
    """
    print("=" * 100)
    print("验证填充后的对账单数据")
    print("=" * 100)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    print(f"\n前{preview_rows}车的数据预览:\n")

    for row_idx in range(2, min(preview_rows + 2, ws.max_row + 1)):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        print(f"车次 {vehicle_no}:")
        print("-" * 80)

        if shop_names_cell:
            shop_lines = str(shop_names_cell).split('\n')
            for i, line in enumerate(shop_lines, 1):
                line = line.strip()
                if not line:
                    continue

                if '-?km' in line:
                    status = "待填充"
                elif re.search(r'-\d+(\.\d+)?km', line):
                    status = "已填充"
                else:
                    status = "无距离"

                print(f"  {i}. {line} [{status}]")

        print()

    # 统计
    print("=" * 100)
    print("统计信息")
    print("=" * 100)

    total_segments = 0
    filled_segments = 0
    missing_segments = 0

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        if shop_names_cell:
            shop_lines = str(shop_names_cell).split('\n')
            for line in shop_lines:
                line = line.strip()
                if line:
                    total_segments += 1
                    if '-?km' in line:
                        missing_segments += 1
                    elif re.search(r'-\d+(\.\d+)?km', line):
                        filled_segments += 1

    print(f"\n总路段数: {total_segments}")
    if total_segments > 0:
        print(f"已填充: {filled_segments} ({filled_segments/total_segments*100:.1f}%)")
        print(f"待填充: {missing_segments} ({missing_segments/total_segments*100:.1f}%)")

    wb.close()


def verify_summary(excel_file):
    """
    生成数据摘要
    """
    print("=" * 100)
    print("对账单数据摘要")
    print("=" * 100)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 按日期统计
    date_stats = defaultdict(lambda: {'vehicles': 0, 'shops': 0})
    total_vehicles = 0
    total_shops = 0

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        total_vehicles += 1
        date_stats[date_value]['vehicles'] += 1

        if shop_names_cell:
            shop_count = len([s for s in str(shop_names_cell).split('\n') if s.strip()])
            total_shops += shop_count
            date_stats[date_value]['shops'] += shop_count

    print(f"\n总车次: {total_vehicles}")
    print(f"总店铺: {total_shops}")
    print(f"日期数: {len(date_stats)}")

    print("\n按日期统计:")
    print("-" * 60)
    for date_val in sorted(date_stats.keys()):
        stats = date_stats[date_val]
        print(f"  {date_val}: {stats['vehicles']} 车, {stats['shops']} 店")

    wb.close()


def verify_complete(excel_file):
    """
    完整性检查
    """
    print("=" * 100)
    print("数据完整性检查")
    print("=" * 100)
    print(f"文件: {excel_file}")

    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    issues = []

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        # 检查必填字段
        if not vehicle_no:
            issues.append(f"行{row_idx}: 缺少车次序号")
        if not date_value:
            issues.append(f"行{row_idx}: 缺少日期")
        if not shop_names_cell:
            issues.append(f"行{row_idx}: 缺少店铺信息")
            continue

        # 检查店铺数据格式
        shop_lines = str(shop_names_cell).split('\n')
        if len(shop_lines) == 0:
            issues.append(f"行{row_idx}: 店铺列表为空")

        for i, line in enumerate(shop_lines, 1):
            line = line.strip()
            if not line:
                continue

            # 检查是否有异常字符
            if '\t' in line:
                issues.append(f"行{row_idx}店铺{i}: 包含制表符")

    if issues:
        print(f"\n发现 {len(issues)} 个问题:")
        for issue in issues[:50]:
            print(f"  - {issue}")
        if len(issues) > 50:
            print(f"  ... 还有 {len(issues) - 50} 个问题未显示")
    else:
        print("\n检查通过！未发现问题。")

    wb.close()


def main():
    parser = argparse.ArgumentParser(description='对账单数据验证工具')
    parser.add_argument('--mode', '-m', required=True,
                        choices=['filled', 'summary', 'complete'],
                        help='验证模式: filled/summary/complete')
    parser.add_argument('--input', '-i', required=True,
                        help='输入Excel文件路径')
    parser.add_argument('--rows', '-r', type=int, default=10,
                        help='预览行数（默认10）')

    args = parser.parse_args()

    if args.mode == 'filled':
        verify_filled_data(args.input, args.rows)
    elif args.mode == 'summary':
        verify_summary(args.input)
    elif args.mode == 'complete':
        verify_complete(args.input)


if __name__ == '__main__':
    main()
