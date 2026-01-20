#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从距离缓存填充距离到对账单Excel
支持合肥和江西两个区域
输出格式：店名-距离km

用法:
    python -m scripts.core.fill_distances --region hefei --input 对账单.xlsx
    python -m scripts.core.fill_distances --region jiangxi --input 对账单.xlsx --output 新对账单.xlsx
"""

import argparse
import os
import re
import sys

import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.utils.common import (
    get_region_config, load_distance_cache, process_route, format_distance
)


def fill_distances_to_excel(region, input_excel, output_excel=None, cache_file=None):
    """
    将距离数据填充到对账单Excel

    Args:
        region: 区域 ('hefei' 或 'jiangxi')
        input_excel: 输入Excel文件路径
        output_excel: 输出Excel文件路径（可选，默认覆盖原文件）
        cache_file: 距离缓存文件路径（可选，默认使用区域配置）
    """
    config = get_region_config(region)
    start_point = config['start_point']

    if cache_file is None:
        cache_file = os.path.join(config['cache_dir'], 'reusable_distances.json')

    if output_excel is None:
        output_excel = input_excel

    print("=" * 60)
    print(f"{region.upper()}仓对账单距离填充")
    print("=" * 60)

    # 读取距离数据
    print(f"\n读取距离数据: {cache_file}")
    distances = load_distance_cache(cache_file)
    print(f"共加载 {len(distances)} 条距离记录")

    # 读取Excel数据
    print(f"\n读取Excel数据: {input_excel}")
    df = pd.read_excel(input_excel, header=None)
    print(f"Excel尺寸: {df.shape}")

    # 使用openpyxl处理以保留格式
    wb = load_workbook(input_excel)
    ws = wb.active

    # 统计信息
    total_routes = 0
    total_stops = 0
    found_count = 0
    not_found_count = 0
    not_found_details = []

    # 处理每一行（从第2行开始，跳过标题行）
    print("\n开始处理路线...")
    print("-" * 60)

    for row_idx in range(2, ws.max_row + 1):
        # 读取店名列（C列，索引3）
        store_cell = ws.cell(row=row_idx, column=3)
        route_text = store_cell.value

        if pd.isna(route_text) or not str(route_text).strip():
            continue

        # 检查是否已经包含距离信息（格式：店名-XXkm）
        first_line = str(route_text).split('\n')[0].strip()
        if re.search(r'-\d+(\.\d+)?km', first_line) or re.search(r'-\?km', first_line):
            print(f"\n第{row_idx}行: 已包含距离信息，跳过")
            continue

        total_routes += 1
        results = process_route(route_text, distances, start_point)

        if not results:
            continue

        # 构建带距离的店名字符串
        formatted_stops = []

        print(f"\n第{row_idx}行 (第{total_routes}车):")
        for r in results:
            total_stops += 1
            if r['found']:
                found_count += 1
                dist_str = format_distance(r['distance'])
                formatted_stops.append(f"{r['stop']}-{dist_str}")
                print(f"  {r['stop']}-{dist_str} [找到]")
            else:
                not_found_count += 1
                formatted_stops.append(f"{r['stop']}-?km")
                not_found_details.append({
                    'row': row_idx,
                    'from': r['from'],
                    'to': r['to']
                })
                print(f"  {r['stop']}-?km [未找到]")

        # 将格式化后的店名（带距离）写回C列
        new_route_text = '\n'.join(formatted_stops)
        ws.cell(row=row_idx, column=3, value=new_route_text)

    # 保存结果
    print("\n" + "=" * 60)
    print("处理完成!")
    print(f"  总路线数: {total_routes}")
    print(f"  总站点数: {total_stops}")
    if total_stops > 0:
        print(f"  找到距离: {found_count} ({found_count/total_stops*100:.1f}%)")
        print(f"  未找到距离: {not_found_count} ({not_found_count/total_stops*100:.1f}%)")

    # 保存文件
    print(f"\n保存结果到: {output_excel}")
    wb.save(output_excel)
    print("保存成功!")

    # 输出未找到的距离详情
    if not_found_details:
        print("\n未找到的距离段:")
        for item in not_found_details[:20]:
            print(f"  行{item['row']}: {item['from']} -> {item['to']}")
        if len(not_found_details) > 20:
            print(f"  ... 还有 {len(not_found_details) - 20} 条未显示")

    return {
        'total_routes': total_routes,
        'total_stops': total_stops,
        'found': found_count,
        'not_found': not_found_count
    }


def main():
    parser = argparse.ArgumentParser(description='从距离缓存填充距离到对账单Excel')
    parser.add_argument('--region', '-r', required=True, choices=['hefei', 'jiangxi'],
                        help='区域: hefei 或 jiangxi')
    parser.add_argument('--input', '-i', required=True,
                        help='输入Excel文件路径')
    parser.add_argument('--output', '-o',
                        help='输出Excel文件路径（可选，默认覆盖原文件）')
    parser.add_argument('--cache', '-c',
                        help='距离缓存文件路径（可选）')

    args = parser.parse_args()

    fill_distances_to_excel(args.region, args.input, args.output, args.cache)


if __name__ == '__main__':
    main()
