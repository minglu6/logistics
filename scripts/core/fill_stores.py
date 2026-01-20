#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从txt文件填充店名到对账单Excel
支持合肥和江西两个区域

用法:
    python -m scripts.core.fill_stores --region hefei --input stores.txt --excel 对账单.xlsx
    python -m scripts.core.fill_stores --region jiangxi --input stores.txt --excel 对账单.xlsx --output 新对账单.xlsx
"""

import argparse
import os
import sys

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from scripts.utils.common import parse_txt_data, date_str_to_excel_serial, get_region_config


def fill_stores_to_excel(input_txt, input_excel, output_excel=None, year=2026):
    """
    将txt中的店名数据填充到Excel

    Args:
        input_txt: 输入txt文件路径
        input_excel: 输入Excel文件路径
        output_excel: 输出Excel文件路径（可选，默认覆盖原文件）
        year: 年份（默认2026）
    """
    if output_excel is None:
        output_excel = input_excel

    print("=" * 60)
    print("填充店名数据到对账单")
    print("=" * 60)

    # 解析txt数据
    print(f"\n读取txt文件: {input_txt}")
    txt_data = parse_txt_data(input_txt)

    print(f"\n解析到 {len(txt_data)} 个日期的数据:")
    for date_str, vehicles in sorted(txt_data.items(), key=lambda x: tuple(map(int, x[0].split('.')))):
        print(f"  {date_str}: {len(vehicles)} 辆车")

    # 加载Excel
    print(f"\n加载Excel文件: {input_excel}")
    wb = openpyxl.load_workbook(input_excel)
    ws = wb.active

    # 找到第一个空白行
    first_empty_row = None
    for i in range(2, 1000):
        if ws[f'B{i}'].value is None and ws[f'C{i}'].value is None:
            first_empty_row = i
            break

    if first_empty_row is None:
        print("未找到空白行!")
        return

    print(f"从第 {first_empty_row} 行开始填充数据")

    # 获取当前最大序号
    current_seq = 0
    for i in range(first_empty_row - 1, 1, -1):
        if ws[f'A{i}'].value is not None:
            current_seq = ws[f'A{i}'].value
            break

    current_row = first_empty_row

    # 按日期排序
    sorted_dates = sorted(txt_data.keys(), key=lambda x: tuple(map(int, x.split('.'))))

    for date_str in sorted_dates:
        vehicles = txt_data[date_str]
        excel_date = date_str_to_excel_serial(date_str, year)

        print(f"\n填充日期: {date_str} (Excel序列号: {excel_date})")
        print(f"  共 {len(vehicles)} 辆车")

        for vehicle_stores in vehicles:
            current_seq += 1

            # 填充序号
            ws[f'A{current_row}'] = current_seq

            # 填充日期
            ws[f'B{current_row}'] = excel_date

            # 填充店名(用换行符连接)
            stores_text = '\n'.join(vehicle_stores)
            ws[f'C{current_row}'] = stores_text

            print(f"  第{current_row}行: 序号={current_seq}, 店铺数={len(vehicle_stores)}")

            current_row += 1

    # 保存文件
    wb.save(output_excel)

    total_rows = current_row - first_empty_row
    print("\n" + "=" * 60)
    print(f"数据填充完成! 共填充 {total_rows} 行数据")
    print(f"保存到: {output_excel}")
    print("=" * 60)

    return total_rows


def main():
    parser = argparse.ArgumentParser(description='从txt文件填充店名到对账单Excel')
    parser.add_argument('--region', '-r', choices=['hefei', 'jiangxi'],
                        help='区域: hefei 或 jiangxi（用于自动路径）')
    parser.add_argument('--input', '-i', required=True,
                        help='输入txt文件路径')
    parser.add_argument('--excel', '-e', required=True,
                        help='输入Excel文件路径')
    parser.add_argument('--output', '-o',
                        help='输出Excel文件路径（可选，默认覆盖原文件）')
    parser.add_argument('--year', '-y', type=int, default=2026,
                        help='年份（默认2026）')

    args = parser.parse_args()

    fill_stores_to_excel(args.input, args.excel, args.output, args.year)


if __name__ == '__main__':
    main()
