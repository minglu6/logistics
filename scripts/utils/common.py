# -*- coding: utf-8 -*-
"""
通用工具函数
包含店名标准化、距离查找、路线解析等常用功能
"""

import re
import json
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from collections import defaultdict

# 区域配置
REGION_CONFIG = {
    'hefei': {
        'start_point': '丰树合肥现代综合产业园',
        'cache_dir': 'data/hefei/cache',
        'summary_dir': 'data/hefei/summary/2026',
        'details_dir': 'data/hefei/details',
    },
    'jiangxi': {
        'start_point': '惠宜选南昌仓',
        'cache_dir': 'data/jiangxi/cache',
        'summary_dir': 'data/jiangxi/summary/2026',
        'details_dir': 'data/jiangxi/details',
    }
}

# 默认冲突阈值（km）
CONFLICT_THRESHOLD = 5.0


def get_region_config(region):
    """获取区域配置"""
    region = region.lower()
    if region not in REGION_CONFIG:
        raise ValueError(f"未知区域: {region}, 支持: {list(REGION_CONFIG.keys())}")
    return REGION_CONFIG[region]


def normalize_store_name(name, aggressive=False):
    """
    标准化店名，处理常见的格式差异

    Args:
        name: 店名
        aggressive: 是否使用激进模式（处理更多变体）

    Returns:
        标准化后的店名，或None如果输入无效
    """
    if pd.isna(name) or name is None:
        return None

    name = str(name).strip()
    if not name:
        return None

    # 统一括号格式
    name = name.replace('（', '(').replace('）', ')')
    # 移除多余空格
    name = re.sub(r'\s+', '', name)

    if aggressive:
        # 更激进的标准化，用于模糊匹配
        # 处理"历臣"和"厉臣"的混用
        name = name.replace('历臣', '厉臣')
        # 处理连字符差异
        name = name.replace('-', '-').replace('—', '-').replace('－', '-')
        # 处理"共橙一站式"和"共橙-站式"的混用
        name = re.sub(r'共橙[-\s]*站式', '共橙一站式', name)
        # 处理"共橙超市"和"供橙超市"
        name = name.replace('供橙超市', '共橙超市')
        name = name.replace('供橙一站式超市', '共橙一站式超市')

    return name


def parse_shop_and_distance(shop_line):
    """
    解析店名行，提取店名和距离

    Args:
        shop_line: 格式如 "店名-XXkm" 或纯店名

    Returns:
        (店名, 距离) 元组，距离可能为None
    """
    shop_line = str(shop_line).strip()
    # 匹配格式：店名-距离km
    pattern = r'^(.+?)-(\d+(?:\.\d+)?)km$'
    match = re.match(pattern, shop_line)

    if match:
        shop_name = match.group(1)
        distance = float(match.group(2))
        return shop_name, distance
    else:
        return shop_line, None


def find_distance(distances, from_store, to_store):
    """
    从距离字典中查找距离，尝试多种格式匹配

    Args:
        distances: 距离字典 {"A -> B": 10.5, ...}
        from_store: 起点店名
        to_store: 终点店名

    Returns:
        (距离值, 匹配的键) 或 (None, None)
    """
    from_norm = normalize_store_name(from_store)
    to_norm = normalize_store_name(to_store)

    if not from_norm or not to_norm:
        return None, None

    # 构建可能的查询键变体
    key_variants = []

    # 原始格式
    key_variants.append(f"{from_store} -> {to_store}")
    key_variants.append(f"{from_norm} -> {to_norm}")

    # 尝试不同的括号格式
    from_half = from_store.replace('（', '(').replace('）', ')')
    to_half = to_store.replace('（', '(').replace('）', ')')
    key_variants.append(f"{from_half} -> {to_half}")

    # 尝试全角括号
    from_full = from_store.replace('(', '（').replace(')', '）')
    to_full = to_store.replace('(', '（').replace(')', '）')
    key_variants.append(f"{from_full} -> {to_full}")

    # 遍历所有变体查找
    for key in key_variants:
        if key in distances:
            return distances[key], key

    # 模糊匹配：使用激进标准化
    from_agg = normalize_store_name(from_store, aggressive=True)
    to_agg = normalize_store_name(to_store, aggressive=True)

    # 遍历所有距离记录进行模糊匹配
    for dist_key, dist_val in distances.items():
        if ' -> ' not in dist_key:
            continue
        parts = dist_key.split(' -> ')
        if len(parts) != 2:
            continue
        key_from, key_to = parts
        key_from_norm = normalize_store_name(key_from)
        key_to_norm = normalize_store_name(key_to)
        key_from_agg = normalize_store_name(key_from, aggressive=True)
        key_to_agg = normalize_store_name(key_to, aggressive=True)

        # 普通匹配
        if key_from_norm == from_norm and key_to_norm == to_norm:
            return dist_val, dist_key

        # 激进匹配
        if key_from_agg == from_agg and key_to_agg == to_agg:
            return dist_val, dist_key

    return None, None


def process_route(route_text, distances, start_point):
    """
    处理一条路线，返回每个站点的距离信息

    Args:
        route_text: 包含换行符的站点列表
        distances: 距离字典
        start_point: 起点名称

    Returns:
        [{'stop': ..., 'from': ..., 'to': ..., 'distance': ..., 'found': ...}, ...]
    """
    if pd.isna(route_text):
        return []

    stops = [s.strip() for s in str(route_text).split('\n') if s.strip()]
    if not stops:
        return []

    results = []

    # 第一站：从起点到第一站
    first_stop = stops[0]
    dist, key = find_distance(distances, start_point, first_stop)
    results.append({
        'stop': first_stop,
        'from': start_point,
        'to': first_stop,
        'distance': dist,
        'key': key,
        'found': dist is not None
    })

    # 后续站点：前一站到下一站
    for i in range(1, len(stops)):
        prev_stop = stops[i-1]
        curr_stop = stops[i]
        dist, key = find_distance(distances, prev_stop, curr_stop)
        results.append({
            'stop': curr_stop,
            'from': prev_stop,
            'to': curr_stop,
            'distance': dist,
            'key': key,
            'found': dist is not None
        })

    return results


def extract_routes_from_excel(file_path, start_point):
    """
    从对账单Excel中提取路线数据

    Args:
        file_path: Excel文件路径
        start_point: 起点名称

    Returns:
        路线列表 [{'vehicle_no': ..., 'date': ..., 'shops': [...], 'distances': [...]}, ...]
    """
    wb = load_workbook(file_path)
    ws = wb.active
    routes = []

    for row_idx in range(2, ws.max_row + 1):
        vehicle_no = ws.cell(row=row_idx, column=1).value
        date_value = ws.cell(row=row_idx, column=2).value
        shop_names_cell = ws.cell(row=row_idx, column=3).value

        if vehicle_no is None:
            break

        shops = []
        distances = []

        if shop_names_cell:
            shop_lines = [s.strip() for s in str(shop_names_cell).split('\n') if s.strip()]

            for shop_line in shop_lines:
                shop_name, distance = parse_shop_and_distance(shop_line)
                shops.append(shop_name)
                distances.append(distance)

        route = {
            'vehicle_no': vehicle_no,
            'date': date_value,
            'shops': shops,
            'distances': distances
        }

        routes.append(route)

    return routes


def build_segments(route, start_point):
    """
    根据物流规则构建路段

    Args:
        route: 路线字典 {'shops': [...], 'distances': [...]}
        start_point: 起点名称

    Returns:
        [(起点, 终点, 距离), ...]
    """
    segments = []
    shops = route['shops']
    distances = route['distances']

    if not shops:
        return segments

    # 第一段：起点到第一站
    first_distance = distances[0] if len(distances) > 0 else None
    segments.append((start_point, shops[0], first_distance))

    # 后续段：前一站到下一站
    for i in range(1, len(shops)):
        distance = distances[i] if i < len(distances) else None
        segments.append((shops[i-1], shops[i], distance))

    return segments


def format_distance(dist_val):
    """
    格式化距离值为字符串

    Args:
        dist_val: 距离数值

    Returns:
        格式化后的字符串，如 "10km" 或 "10.5km"
    """
    if dist_val is None:
        return "?km"
    if dist_val == int(dist_val):
        return f"{int(dist_val)}km"
    return f"{dist_val}km"


def load_distance_cache(cache_file):
    """加载距离缓存"""
    try:
        with open(cache_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return {}


def save_distance_cache(cache_file, distances):
    """保存距离缓存"""
    with open(cache_file, 'w', encoding='utf-8') as f:
        json.dump(distances, f, ensure_ascii=False, indent=2)


def date_str_to_excel_serial(date_str, year=2026):
    """
    将日期字符串(如"1.13")转换为Excel日期序列号

    Args:
        date_str: 日期字符串，格式如 "1.13"
        year: 年份，默认2026

    Returns:
        Excel日期序列号
    """
    parts = date_str.split('.')
    month = int(parts[0])
    day = int(parts[1])

    date_obj = datetime(year, month, day)
    excel_epoch = datetime(1899, 12, 30)
    delta = date_obj - excel_epoch

    return delta.days


def parse_txt_data(txt_file):
    """
    解析txt文件，返回按日期分组的店名数据

    Args:
        txt_file: txt文件路径

    Returns:
        {日期: [[车1店名], [车2店名], ...], ...}
    """
    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    data = {}
    current_date = None
    current_vehicle = []

    for line in lines:
        line = line.strip()

        # 跳过空行
        if not line:
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []
            continue

        # 检查是否是日期行(格式如"1.13")
        if line.replace('.', '').isdigit() and '.' in line:
            # 保存前一辆车的数据
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []
            current_date = line
        else:
            # 店名行
            current_vehicle.append(line)

    # 保存最后一辆车的数据
    if current_vehicle and current_date:
        if current_date not in data:
            data[current_date] = []
        data[current_date].append(current_vehicle)

    return data


def extract_stores_from_excel(excel_file, store_column=4):
    """
    从Excel文件中提取店名，按车辆分组

    Args:
        excel_file: Excel文件路径
        store_column: 店名所在列（默认第4列）

    Returns:
        [[车1的店名列表], [车2的店名列表], ...]
    """
    wb = load_workbook(excel_file)
    ws = wb.active

    vehicles = []
    current_vehicle = []

    for row_idx in range(2, ws.max_row + 1):
        store_name = ws.cell(row=row_idx, column=store_column).value

        if store_name is None or str(store_name).strip() == '':
            if current_vehicle:
                vehicles.append(current_vehicle)
                current_vehicle = []
        else:
            current_vehicle.append(str(store_name).strip())

    if current_vehicle:
        vehicles.append(current_vehicle)

    return vehicles
