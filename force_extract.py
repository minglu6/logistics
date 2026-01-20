#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import sys
sys.stdout = open('extraction_log.txt', 'w', encoding='utf-8')
sys.stderr = sys.stdout

try:
    print("="*60)
    print("开始提取合肥未解决物流店名数据")
    print("="*60)
    
    import openpyxl
    print("openpyxl模块加载成功")
    
    import os
    print(f"当前工作目录: {os.getcwd()}")
    
    base_dir = 'data/hefei/details/unresolved'
    output_file = 'data/hefei/details/unresolved/物流店名数据_未解决.txt'
    
    files = [
        ('1.13', '临努1.13.xlsx'),
        ('1.15', '临努1.15.xlsx'),
        ('1.16', '临努1.16.xlsx'),
        ('1.17', '临努1.17.xlsx'),
        ('1.18', '临努1.18.xlsx'),
        ('1.19', '临努1.19.xlsx'),
    ]
    
    all_data = {}
    
    for date_str, filename in files:
        file_path = os.path.join(base_dir, filename)
        print(f"\n处理文件: {date_str} - {file_path}")
        
        if not os.path.exists(file_path):
            print(f"  警告: 文件不存在")
            continue
        
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        print(f"  工作表已加载，总行数: {ws.max_row}")
        
        vehicles = []
        current_vehicle = []
        
        for row_idx in range(2, ws.max_row + 1):
            store_name = ws.cell(row=row_idx, column=4).value
            
            if store_name is None or str(store_name).strip() == '':
                if current_vehicle:
                    vehicles.append(current_vehicle)
                    current_vehicle = []
            else:
                current_vehicle.append(str(store_name).strip())
        
        if current_vehicle:
            vehicles.append(current_vehicle)
        
        all_data[date_str] = vehicles
        print(f"  提取到 {len(vehicles)} 辆车的数据")
        for i, v in enumerate(vehicles, 1):
            print(f"    第{i}车: {len(v)}个店铺")
    
    print(f"\n写入输出文件: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        for date_str in ['1.13', '1.15', '1.16', '1.17', '1.18', '1.19']:
            if date_str not in all_data:
                continue
            
            f.write(f"{date_str}\n\n")
            
            vehicles = all_data[date_str]
            for vehicle_stores in vehicles:
                for store in vehicle_stores:
                    f.write(f"{store}\n")
                f.write("\n")
    
    print("="*60)
    print("提取完成!")
    print(f"输出文件: {output_file}")
    print("="*60)
    
except Exception as e:
    print(f"\n错误: {e}")
    import traceback
    traceback.print_exc()
finally:
    sys.stdout.close()
