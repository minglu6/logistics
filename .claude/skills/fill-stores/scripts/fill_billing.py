#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
通用对账单填充脚本
支持合肥、江西等区域的对账单数据填充
从txt文件读取物流店名数据,填充到Excel表格中
"""
import openpyxl
from datetime import datetime
import sys
import os
import shutil


def parse_txt_data(txt_file):
    """
    解析txt文件,返回按日期分组的店名数据
    格式: {日期: [车辆列表]}
    """
    with open(txt_file, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    data = {}
    current_date = None
    current_vehicle = []

    for line in lines:
        line = line.strip()

        # 跳过空行
        if not line:
            # 如果当前有车辆数据,保存它
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []
            continue

        # 检查是否是日期行(格式如"1.9")
        if line.replace('.', '').isdigit() and '.' in line:
            # 保存前一辆车的数据
            if current_vehicle and current_date:
                if current_date not in data:
                    data[current_date] = []
                data[current_date].append(current_vehicle)
                current_vehicle = []

            current_date = line
        else:
            # 店名行
            current_vehicle.append(line)

    # 保存最后一辆车的数据
    if current_vehicle and current_date:
        if current_date not in data:
            data[current_date] = []
        data[current_date].append(current_vehicle)

    return data


def date_str_to_excel_serial(date_str, year=2026):
    """
    将日期字符串(如"1.9")转换为Excel日期序列号
    """
    parts = date_str.split('.')
    month = int(parts[0])
    day = int(parts[1])

    # 默认2026年
    date_obj = datetime(year, month, day)

    # Excel的日期序列号从1899-12-30开始
    excel_epoch = datetime(1899, 12, 30)
    delta = date_obj - excel_epoch

    return delta.days


def fill_excel_data(source_excel, txt_data, output_excel, year=2026):
    """
    将txt数据填充到Excel文件中
    """
    # 如果输出文件不是源文件，先复制
    if source_excel != output_excel:
        print(f"复制文件: {source_excel} -> {output_excel}")
        shutil.copy2(source_excel, output_excel)

    wb = openpyxl.load_workbook(output_excel)
    ws = wb.active

    # 找到第一个空白行
    first_empty_row = None
    for i in range(2, 500):
        if ws[f'B{i}'].value is None and ws[f'C{i}'].value is None:
            first_empty_row = i
            break

    if first_empty_row is None:
        print("未找到空白行!")
        return

    print(f"从第 {first_empty_row} 行开始填充数据")

    # 获取当前最大序号
    current_seq = ws[f'A{first_empty_row - 1}'].value
    if current_seq is None:
        # 向上查找最后一个有序号的行
        for i in range(first_empty_row - 1, 1, -1):
            if ws[f'A{i}'].value is not None:
                current_seq = ws[f'A{i}'].value
                break
        if current_seq is None:
            current_seq = 0

    current_row = first_empty_row

    # 按日期排序
    sorted_dates = sorted(txt_data.keys(), key=lambda x: tuple(map(int, x.split('.'))))

    for date_str in sorted_dates:
        vehicles = txt_data[date_str]
        excel_date = date_str_to_excel_serial(date_str, year)

        print(f"\n填充日期: {date_str} (Excel序列号: {excel_date})")
        print(f"  共 {len(vehicles)} 辆车")

        for vehicle_stores in vehicles:
            current_seq += 1

            # 填充序号
            ws[f'A{current_row}'] = current_seq

            # 填充日期
            ws[f'B{current_row}'] = excel_date

            # 填充店名(用换行符连接)
            stores_text = '\n'.join(vehicle_stores)
            ws[f'C{current_row}'] = stores_text

            print(f"  第{current_row}行: 序号={current_seq}, 店铺数={len(vehicle_stores)}, 首店={vehicle_stores[0][:30]}...")

            current_row += 1

    # 保存文件
    wb.save(output_excel)
    print(f"\n数据填充完成! 共填充 {current_row - first_empty_row} 行数据")
    print(f"保存到: {output_excel}")


def extract_date_suffix(filename):
    """
    从文件名中提取日期后缀
    例如: 惠宜选合肥仓1月份对账单0112.xlsx -> 0112
    """
    import re
    match = re.search(r'(\d{4})\.xlsx$', filename)
    if match:
        return match.group(1)
    return None


def generate_output_filename(source_excel, new_date_suffix):
    """
    根据源文件名和新日期后缀生成输出文件名
    例如: 惠宜选合肥仓1月份对账单0112.xlsx + 0119 -> 惠宜选合肥仓1月份对账单0119.xlsx
    """
    import re
    # 替换日期后缀
    output_name = re.sub(r'\d{4}\.xlsx$', f'{new_date_suffix}.xlsx', source_excel)
    return output_name


def main():
    if len(sys.argv) < 4:
        print("Usage: python fill_billing.py <source_excel> <stores_txt> <date_suffix> [year]")
        print()
        print("Parameters:")
        print("  source_excel  : 源对账单Excel文件 (例如: data/hefei/summary/2026/惠宜选合肥仓1月份对账单0112.xlsx)")
        print("  stores_txt    : 店名数据txt文件 (例如: data/hefei/details/unresolved/物流店名数据_1.13_1.19.txt)")
        print("  date_suffix   : 新的日期后缀 (例如: 0119)")
        print("  year          : 年份 (可选, 默认2026)")
        print()
        print("Examples:")
        print("  # 合肥")
        print("  python fill_billing.py data/hefei/summary/2026/惠宜选合肥仓1月份对账单0112.xlsx \\")
        print("                         data/hefei/details/unresolved/物流店名数据_1.13_1.19.txt \\")
        print("                         0119")
        print()
        print("  # 江西")
        print("  python fill_billing.py data/jiangxi/summary/2026/江西对账单0110.xlsx \\")
        print("                         data/jiangxi/details/物流店名数据_1.11_1.13.txt \\")
        print("                         0113")
        sys.exit(1)

    source_excel = sys.argv[1]
    stores_txt = sys.argv[2]
    date_suffix = sys.argv[3]
    year = int(sys.argv[4]) if len(sys.argv) > 4 else 2026

    # 验证文件存在
    if not os.path.exists(source_excel):
        print(f"Error: 源Excel文件不存在 - {source_excel}")
        sys.exit(1)

    if not os.path.exists(stores_txt):
        print(f"Error: 店名txt文件不存在 - {stores_txt}")
        sys.exit(1)

    # 生成输出文件名
    output_excel = generate_output_filename(source_excel, date_suffix)

    print("=" * 70)
    print("通用对账单填充工具")
    print("=" * 70)
    print(f"\n源Excel文件: {source_excel}")
    print(f"店名txt文件: {stores_txt}")
    print(f"输出Excel文件: {output_excel}")
    print(f"年份: {year}")
    print()

    # 解析txt数据
    print("解析店名数据...")
    txt_data = parse_txt_data(stores_txt)

    print(f"\n解析到 {len(txt_data)} 个日期的数据:")
    for date_str, vehicles in sorted(txt_data.items(), key=lambda x: tuple(map(int, x[0].split('.')))):
        print(f"  {date_str}: {len(vehicles)} 辆车")

    # 填充Excel
    print("\n开始填充数据...")
    fill_excel_data(source_excel, txt_data, output_excel, year)

    print("\n" + "=" * 70)
    print("处理完成!")
    print("=" * 70)


if __name__ == '__main__':
    main()
