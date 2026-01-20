#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
更新距离缓存脚本
从对账单Excel文件中提取距离数据，更新对应区域的缓存文件
支持合肥和江西两个仓库
"""

import json
import re
import sys
import os
from openpyxl import load_workbook


def detect_region(excel_path):
    """
    根据文件路径自动检测区域
    """
    excel_path_lower = excel_path.lower()
    if 'hefei' in excel_path_lower or '合肥' in excel_path_lower:
        return 'hefei'
    elif 'jiangxi' in excel_path_lower or '江西' in excel_path_lower:
        return 'jiangxi'
    return None


def get_cache_path(region):
    """
    获取区域对应的缓存文件路径
    """
    if region == 'hefei':
        return 'data/hefei/cache/reusable_distances.json'
    elif region == 'jiangxi':
        return 'data/jiangxi/cache/reusable_distances.json'
    return None


def extract_distances_from_route(route_text, start_point=None):
    """
    从路线文本中提取店名和距离的配对
    返回: [(from_store, to_store, distance), ...]
    """
    if not route_text:
        return []

    stops = [s.strip() for s in str(route_text).split('\n') if s.strip()]
    distances = []

    for i in range(len(stops)):
        # 提取店名和距离 (格式: 店名-XXkm)
        match = re.match(r'(.+?)-(\d+(?:\.\d+)?)km', stops[i])
        if not match:
            continue

        store_name = match.group(1)
        dist = float(match.group(2))

        # 第一站：从起点到第一个店
        if i == 0 and start_point:
            distances.append((start_point, store_name, dist))

        # 后续站：从前一站到当前站
        if i > 0:
            prev_match = re.match(r'(.+?)-\d+(?:\.\d+)?km', stops[i-1])
            if prev_match:
                from_store = prev_match.group(1)
                distances.append((from_store, store_name, dist))

    return distances


def extract_all_distances_from_excel(excel_path, start_point=None):
    """
    从Excel文件中提取所有距离数据
    返回: {key: distance}字典
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    all_distances = {}
    routes_processed = 0

    # 遍历所有行（从第2行开始，跳过标题行）
    for row_idx in range(2, ws.max_row + 1):
        # 读取店名列（C列，索引3）
        store_cell = ws.cell(row=row_idx, column=3)
        route_text = store_cell.value

        if route_text is None or not str(route_text).strip():
            continue

        # 检查是否包含距离信息
        first_line = str(route_text).split('\n')[0].strip()
        if not re.search(r'-\d+(\.\d+)?km', first_line):
            continue

        routes_processed += 1

        # 提取这条路线的所有距离配对
        route_distances = extract_distances_from_route(route_text, start_point)

        for from_store, to_store, dist in route_distances:
            key = f"{from_store} -> {to_store}"
            # 如果同一个key有多个距离值，保留第一个（或者可以选择保留最小值）
            if key not in all_distances:
                all_distances[key] = dist

    wb.close()
    return all_distances, routes_processed


def load_cache(cache_path):
    """
    加载现有缓存，如果文件不存在则返回空字典
    """
    if os.path.exists(cache_path):
        with open(cache_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    else:
        # 如果缓存文件不存在，创建目录
        os.makedirs(os.path.dirname(cache_path), exist_ok=True)
        return {}


def save_cache(cache_path, cache_data):
    """
    保存缓存到文件
    """
    with open(cache_path, 'w', encoding='utf-8') as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)


def main():
    if len(sys.argv) < 2:
        print("Usage: python update_cache.py <excel_file> [region]")
        print()
        print("Parameters:")
        print("  excel_file : 对账单Excel文件路径（已填充距离信息的）")
        print("  region     : 区域 (hefei/jiangxi，可选，默认自动检测)")
        print()
        print("Examples:")
        print("  # 自动检测区域")
        print("  python update_cache.py data/hefei/summary/2026/惠宜选合肥仓1月份对账单0119.xlsx")
        print()
        print("  # 指定区域")
        print("  python update_cache.py data/jiangxi/summary/2026/江西对账单0113.xlsx jiangxi")
        sys.exit(1)

    excel_path = sys.argv[1]

    # 验证Excel文件存在
    if not os.path.exists(excel_path):
        print(f"Error: Excel文件不存在 - {excel_path}")
        sys.exit(1)

    # 确定区域
    if len(sys.argv) > 2:
        region = sys.argv[2].lower()
    else:
        region = detect_region(excel_path)

    if region not in ['hefei', 'jiangxi']:
        print(f"Error: 无法识别区域。请指定 region 参数 (hefei/jiangxi)")
        print(f"检测到的区域: {region}")
        sys.exit(1)

    # 获取缓存路径
    cache_path = get_cache_path(region)
    if cache_path is None:
        print(f"Error: 无法确定缓存路径")
        sys.exit(1)

    # 各仓库的起点
    start_points = {
        'hefei': "丰树合肥现代综合产业园",
        'jiangxi': "南昌红谷滩区"
    }
    start_point = start_points.get(region)

    print("=" * 70)
    print("距离缓存更新工具")
    print("=" * 70)
    print(f"\n区域: {region.upper()}")
    print(f"起点: {start_point}")
    print(f"Excel文件: {excel_path}")
    print(f"缓存文件: {cache_path}")

    # 加载现有缓存
    print(f"\n读取现有缓存...")
    cache_data = load_cache(cache_path)
    original_count = len(cache_data)
    print(f"当前缓存中有 {original_count} 条距离记录")

    # 从Excel提取距离数据
    print(f"\n从Excel提取距离数据...")
    excel_distances, routes_processed = extract_all_distances_from_excel(excel_path, start_point)
    print(f"从 {routes_processed} 条路线中提取到 {len(excel_distances)} 条距离配对")

    # 找出新的距离数据
    new_distances = {}
    updated_distances = {}

    for key, new_dist in excel_distances.items():
        if key not in cache_data:
            new_distances[key] = new_dist
        elif cache_data[key] != new_dist:
            # 距离值不同，记录为更新
            updated_distances[key] = {
                'old': cache_data[key],
                'new': new_dist
            }

    # 显示统计
    print("\n" + "=" * 70)
    print("分析结果:")
    print(f"  Excel中的距离配对: {len(excel_distances)}")
    print(f"  缓存中已存在: {len(excel_distances) - len(new_distances) - len(updated_distances)}")
    print(f"  需要新增: {len(new_distances)}")
    print(f"  需要更新: {len(updated_distances)}")

    # 显示新增的距离
    if new_distances:
        print("\n新增的距离数据:")
        print("-" * 70)
        for key, dist in sorted(new_distances.items())[:20]:  # 只显示前20条
            print(f'  "{key}": {dist}')
        if len(new_distances) > 20:
            print(f"  ... 还有 {len(new_distances) - 20} 条")

    # 显示需要更新的距离
    if updated_distances:
        print("\n距离值发生变化（将保留原值）:")
        print("-" * 70)
        for key, vals in sorted(updated_distances.items())[:10]:  # 只显示前10条
            print(f'  "{key}":')
            print(f'    缓存值: {vals["old"]} km')
            print(f'    Excel值: {vals["new"]} km')
        if len(updated_distances) > 10:
            print(f"  ... 还有 {len(updated_distances) - 10} 条")

    # 更新缓存
    if new_distances:
        cache_data.update(new_distances)
        save_cache(cache_path, cache_data)

        print("\n" + "=" * 70)
        print("缓存更新完成!")
        print(f"  原有记录: {original_count}")
        print(f"  新增记录: {len(new_distances)}")
        print(f"  更新后总数: {len(cache_data)}")
        print(f"  缓存文件: {cache_path}")
        print("=" * 70)
    else:
        print("\n" + "=" * 70)
        print("没有新的距离数据需要添加")
        print("=" * 70)

    # 如果有冲突的距离值，给出提示
    if updated_distances:
        print("\n注意: 发现 {} 条距离配对的值与缓存不同，已保留缓存中的原值。".format(len(updated_distances)))
        print("如需更新这些值，请手动修改缓存文件。")


if __name__ == '__main__':
    main()
