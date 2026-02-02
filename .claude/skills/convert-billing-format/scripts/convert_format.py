#!/usr/bin/env python3
"""
将简化格式（临努格式）的对账单转换为标准模板格式（江西仓对账单模板格式）

输入格式：日期 | 区域 | 门店（店名：公里数，店名：公里数）| 公里数 | 单价 | 运费 | 备注
输出格式：序号 | 日期 | 店名 | 公里数 | 公里数 | 不含税单价 | 含税单价 | 不含税合价 | 含税合价 | 司机价格 | 司机价格 | 司机姓名 | 照片 | 备注
"""

import sys
import re
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment


def parse_stores(stores_raw: str) -> str:
    """
    将门店字符串从"店名1：公里数1，店名2：公里数2"格式
    转换为"店名1-公里数1km\n店名2-公里数2km"格式
    """
    if not stores_raw:
        return ""

    stores_list = stores_raw.split('，')
    store_entries = []

    for store in stores_list:
        # 支持中文冒号和英文冒号
        parts = re.split(r'[:：]', store)
        if len(parts) >= 2:
            store_name = parts[0].strip()
            km = parts[1].strip()
            store_entries.append(f'{store_name}-{km}km')
        elif parts[0].strip():
            store_entries.append(parts[0].strip())

    return '\n'.join(store_entries)


def convert_billing_format(source_path: str, output_path: str = None):
    """
    转换对账单格式

    Args:
        source_path: 源文件路径
        output_path: 输出文件路径，默认在源文件名后添加_转换后
    """
    # 确定输出路径
    if output_path is None:
        base, ext = os.path.splitext(source_path)
        output_path = f"{base}_转换后{ext}"

    # 读取源文件
    wb_source = load_workbook(source_path, data_only=True)
    ws_source = wb_source.active

    # 创建新工作簿
    wb_new = Workbook()
    ws_new = wb_new.active

    # 定义对齐样式
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 写入表头
    headers = [
        '序号', '日期', '店名', '公里数', '公里数',
        '不含税单价', '含税单价', '不含税合价（运费）', '含税合价（运费）',
        '司机价格', '司机价格', '司机姓名', '照片', '备注'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_new.cell(row=1, column=col, value=header)
        cell.alignment = align_center

    # 处理数据行
    row_num = 2
    seq_num = 1

    for row in ws_source.iter_rows(min_row=2, values_only=True):
        if row[0] is None:  # 跳过空行
            continue

        date_val = row[0]      # 日期（第1列）
        stores_raw = row[2]    # 门店（第3列）
        km_total = row[3]      # 总公里数（第4列）
        remark = row[6] if len(row) > 6 else None  # 备注（第7列）

        if stores_raw is None:
            continue

        # 转换门店格式
        stores_formatted = parse_stores(stores_raw)

        r = row_num

        # 序号
        cell = ws_new.cell(row=r, column=1, value=seq_num)
        cell.alignment = align_center

        # 日期 - 设置为"1月1日"格式
        cell = ws_new.cell(row=r, column=2, value=date_val)
        cell.alignment = align_center
        cell.number_format = 'M"月"D"日"'

        # 店名
        cell = ws_new.cell(row=r, column=3, value=stores_formatted)
        cell.alignment = align_left

        # 公里数（第4、5列）
        cell = ws_new.cell(row=r, column=4, value=km_total)
        cell.alignment = align_center

        cell = ws_new.cell(row=r, column=5, value=km_total)
        cell.alignment = align_center

        # 公式列
        # 不含税单价
        ws_new.cell(row=r, column=6, value=f'=G{r}/1.09').alignment = align_center
        # 含税单价
        ws_new.cell(row=r, column=7, value=f'=IF(D{r}<=100,440,IF(D{r}<=200,4.2,IF(D{r}<=300,4,3.9)))').alignment = align_center
        # 不含税合价
        ws_new.cell(row=r, column=8, value=f'=D{r}*F{r}').alignment = align_center
        # 含税合价
        ws_new.cell(row=r, column=9, value=f'=D{r}*G{r}').alignment = align_center
        # 司机价格
        ws_new.cell(row=r, column=10, value=f'=IF(D{r}<=100,400,IF(D{r}<=300,3.2,3))').alignment = align_center
        # 司机合价
        ws_new.cell(row=r, column=11, value=f'=D{r}*J{r}').alignment = align_center

        # 备注
        if remark:
            cell = ws_new.cell(row=r, column=14, value=remark)
            cell.alignment = align_left

        row_num += 1
        seq_num += 1

    # 调整列宽
    ws_new.column_dimensions['C'].width = 50  # 店名列
    ws_new.column_dimensions['B'].width = 10  # 日期列

    # 保存文件
    wb_new.save(output_path)

    print(f"转换完成!")
    print(f"源文件: {source_path}")
    print(f"输出文件: {output_path}")
    print(f"共转换 {seq_num - 1} 行数据")


def main():
    if len(sys.argv) < 2:
        print("用法: python convert_format.py <源文件> [输出文件]")
        print("示例: python convert_format.py '惠宜选物流对账单--临努--1月(2).xlsx'")
        sys.exit(1)

    source_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None

    if not os.path.exists(source_path):
        print(f"错误: 源文件不存在: {source_path}")
        sys.exit(1)

    convert_billing_format(source_path, output_path)


if __name__ == "__main__":
    main()
