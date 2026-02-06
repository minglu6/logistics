#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
从合肥物流 Excel 文件中提取店名数据
支持自动检测日期范围并生成输出文件名
"""
import openpyxl
import os
import sys
import re
from pathlib import Path


def extract_date_from_filename(filename):
    """
    从文件名中提取日期
    例如：临努1.13.xlsx -> 1.13
    """
    match = re.search(r'(\d+\.\d+)', filename)
    if match:
        return match.group(1)
    return None


def extract_stores_from_excel(excel_file):
    """
    从Excel文件中提取店名，按车辆分组
    返回：[[车1的店名列表], [车2的店名列表], ...]
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    vehicles = []
    current_vehicle = []

    # 从第2行开始读取（跳过标题行）
    for row_idx in range(2, ws.max_row + 1):
        # 读取店名（第4列）
        store_name = ws.cell(row=row_idx, column=4).value

        # 如果是空行，表示一车结束
        if store_name is None or str(store_name).strip() == '':
            if current_vehicle:
                vehicles.append(current_vehicle)
                current_vehicle = []
        else:
            current_vehicle.append(str(store_name).strip())

    # 保存最后一车
    if current_vehicle:
        vehicles.append(current_vehicle)

    return vehicles


def sort_dates(dates):
    """
    对日期字符串进行排序
    例如：['1.9', '1.13', '1.2'] -> ['1.2', '1.9', '1.13']
    """
    def date_key(date_str):
        parts = date_str.split('.')
        if len(parts) == 2:
            return (int(parts[0]), int(parts[1]))
        return (0, 0)

    return sorted(dates, key=date_key)


def main():
    if len(sys.argv) < 2:
        print("Usage: python extract_stores.py <input_directory> [output_file]")
        print()
        print("Examples:")
        print("  python extract_stores.py data/hefei/details/unresolved")
        print("  python extract_stores.py data/hefei/details/unresolved output.txt")
        sys.exit(1)

    # 获取输入目录
    input_dir = sys.argv[1]
    if not os.path.exists(input_dir):
        print(f"Error: Directory not found - {input_dir}")
        sys.exit(1)

    # 扫描目录中的所有 Excel 文件
    print("=" * 60)
    print("提取合肥物流店名数据")
    print("=" * 60)
    print(f"\n扫描目录: {input_dir}")

    excel_files = []
    for filename in os.listdir(input_dir):
        if filename.endswith('.xlsx') and (filename.startswith('临努') or re.match(r'^\d+\.\d+\.xlsx$', filename)):
            date_str = extract_date_from_filename(filename)
            if date_str:
                file_path = os.path.join(input_dir, filename)
                excel_files.append((date_str, file_path, filename))

    if not excel_files:
        print(f"Error: No Excel files found matching pattern '临努*.xlsx' or '日期.xlsx' in {input_dir}")
        sys.exit(1)

    # 按日期排序
    excel_files.sort(key=lambda x: (int(x[0].split('.')[0]), int(x[0].split('.')[1])))

    print(f"找到 {len(excel_files)} 个文件\n")

    # 提取所有数据
    all_data = {}
    date_list = []

    for date_str, file_path, filename in excel_files:
        print(f"处理 {date_str}: {filename}")
        vehicles = extract_stores_from_excel(file_path)
        all_data[date_str] = vehicles
        date_list.append(date_str)
        print(f"  提取到 {len(vehicles)} 辆车")
        for i, vehicle in enumerate(vehicles, 1):
            if vehicle:
                print(f"    第{i}车: {len(vehicle)}个店铺, 首店: {vehicle[0][:30]}...")

    # 确定输出文件
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # 自动生成输出文件名
        sorted_dates = sort_dates(date_list)
        if sorted_dates:
            start_date = sorted_dates[0]
            end_date = sorted_dates[-1]
            output_filename = f"物流店名数据_{start_date}_{end_date}.txt"
            output_file = os.path.join(input_dir, output_filename)
        else:
            output_file = os.path.join(input_dir, "物流店名数据.txt")

    # 写入输出文件
    print(f"\n写入输出文件: {output_file}")

    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_file) if os.path.dirname(output_file) else '.', exist_ok=True)

    with open(output_file, 'w', encoding='utf-8') as f:
        sorted_dates = sort_dates(date_list)
        for date_str in sorted_dates:
            if date_str not in all_data:
                continue

            # 写入日期标题
            f.write(f"{date_str}\n")
            f.write("\n")

            # 写入该日期的所有车辆
            vehicles = all_data[date_str]
            for vehicle_stores in vehicles:
                for store in vehicle_stores:
                    f.write(f"{store}\n")
                f.write("\n")  # 车辆之间用空行分隔

    print("=" * 60)
    print("提取完成!")
    print(f"输出文件: {output_file}")
    print(f"总计: {len(date_list)} 个日期, {sum(len(all_data[d]) for d in date_list)} 辆车")
    print("=" * 60)


if __name__ == '__main__':
    main()
