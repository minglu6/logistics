import openpyxl
import os

base_dir = 'data/hefei/details/unresolved'
output_file = 'data/hefei/details/unresolved/stores_unresolved.txt'

files = [
    ('1.13', os.path.join(base_dir, '临努1.13.xlsx')),
]

all_stores = []

for date_str, file_path in files:
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        for row_idx in range(2, ws.max_row + 1):
            store_name = ws.cell(row=row_idx, column=4).value
            if store_name and str(store_name).strip():
                all_stores.append(str(store_name).strip())

with open(output_file, 'w', encoding='utf-8') as f:
    for store in all_stores:
        f.write(f'{store}\n')
