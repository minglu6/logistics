import openpyxl

# File paths
linnu_path = "data/hefei/final/合肥仓物流对账-临努-12月份.xlsx"
ref_path = "data/hefei/summary/2025/惠宜选合肥仓12月份对账单.xlsx"

# Load workbooks
wb_linnu = openpyxl.load_workbook(linnu_path, data_only=True)
wb_ref = openpyxl.load_workbook(ref_path, data_only=True)

ws_linnu = wb_linnu.active
ws_ref = wb_ref.active

rows_to_check = [12, 77, 98]

for row_num in rows_to_check:
    print("=" * 80)
    print(f"ROW {row_num}")
    print("=" * 80)

    # Get raw cell values from Column C
    linnu_val = ws_linnu.cell(row=row_num, column=3).value
    ref_val = ws_ref.cell(row=row_num, column=3).value

    print(f"\n--- 临努 file (Column C, row {row_num}) ---")
    print(f"Raw value: {repr(linnu_val)}")

    print(f"\n--- Reference file (Column C, row {row_num}) ---")
    print(f"Raw value: {repr(ref_val)}")

    # Parse into individual stores
    if linnu_val is not None:
        linnu_stores = [s.strip() for s in str(linnu_val).split('\uff0c') if s.strip()]
    else:
        linnu_stores = []

    if ref_val is not None:
        ref_stores = [s.strip() for s in str(ref_val).split('\n') if s.strip()]
    else:
        ref_stores = []

    print(f"\n--- Parsed stores comparison (临努: {len(linnu_stores)} stores, Reference: {len(ref_stores)} stores) ---")
    max_len = max(len(linnu_stores), len(ref_stores))
    print(f"{'#':<4} {'临努 file':<40} {'Reference file':<40} {'Match?'}")
    print("-" * 90)
    for i in range(max_len):
        ls = linnu_stores[i] if i < len(linnu_stores) else "<MISSING>"
        rs = ref_stores[i] if i < len(ref_stores) else "<MISSING>"
        match = "OK" if ls == rs else "MISMATCH"
        print(f"{i+1:<4} {ls:<40} {rs:<40} {match}")

    # Check for stores in one but not the other
    linnu_set = set(linnu_stores)
    ref_set = set(ref_stores)

    only_in_linnu = linnu_set - ref_set
    only_in_ref = ref_set - linnu_set

    if only_in_linnu:
        print(f"\n  Stores ONLY in 临努 file: {only_in_linnu}")
    if only_in_ref:
        print(f"\n  Stores ONLY in Reference file: {only_in_ref}")
    if not only_in_linnu and not only_in_ref:
        print(f"\n  Same stores in both files (but possibly different order)")

    print()

