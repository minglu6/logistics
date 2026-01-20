import openpyxl
import os
import traceback

try:
    file_path = 'data/hefei/details/unresolved/临努1.13.xlsx'
    output_path = 'data/hefei/details/unresolved/debug_output.txt'
    
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f'开始处理文件: {file_path}\n')
        f.write(f'文件存在: {os.path.exists(file_path)}\n')
        
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            f.write(f'工作表已加载，总行数: {ws.max_row}\n\n')
            
            # 提取前10行数据
            for row_idx in range(2, min(12, ws.max_row + 1)):
                store_name = ws.cell(row=row_idx, column=4).value
                f.write(f'第{row_idx}行第4列: {store_name}\n')
                
        f.write('\n处理完成\n')
        
except Exception as e:
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f'错误: {str(e)}\n')
        f.write(traceback.format_exc())
